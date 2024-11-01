from openpyxl import load_workbook
import time
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import requests
from selenium.webdriver.common.keys import Keys  # Para acessar as teclas especiais
from selenium.webdriver.common.action_chains import ActionChains
import pyautogui
import undetected_chromedriver as uc



caminhoExcel = r'C:\Users\diogo.lana\Desktop\PYTHON\Projetos Dukar\teste.xlsx'

#Abri a planilha do Excel
planilha = load_workbook(caminhoExcel)
# Acessa as guias pelo nome
guia_relacao_veiculos = planilha['RELACAO DE VEICULOS']
guia_resultado_autuacao = planilha['RESULTADO AUTUACAO']
guia_resultado_multa = planilha['RESULTADO MULTA']


# Inicializa a variável linhaDaVez antes do loop
cont = 1

# Configurar o Chrome com um User-Agent falso usando undetected-chromedriver
chrome_options = Options()
chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36")

# Inicializa o navegador com as opções e com o stealth mode ativado
navegador = uc.Chrome(options=chrome_options)
navegador.maximize_window()

# Sua chave API do 2Captcha
API_KEY = '657c1d808a967e254d096cd0cfd696c3'

# data-sitekey do reCAPTCHA (obtido ao inspecionar a página)
GOOGLE_KEY = '6LfVpnIUAAAAAHkISk6Z6juZcsUx6hbyJGwfnfPL'

# URL da página com o reCAPTCHA
PAGE_URL = 'https://www.transito.mg.gov.br/veiculos/situacao-do-veiculo/consultar-situacao-do-veiculo/'

# 1. Enviar requisição para o 2Captcha para resolver o reCAPTCHA
def enviar_requisicao_captcha(api_key, google_key, page_url):
    url = f"http://2captcha.com/in.php?key={api_key}&method=userrecaptcha&googlekey={google_key}&pageurl={page_url}"
    response = requests.get(url)
    if response.status_code == 200 and 'OK|' in response.text:
        captcha_id = response.text.split('|')[1]
        return captcha_id
    else:
        raise Exception(f"Erro ao enviar captcha: {response.text}")

# 2. Obter o token de resposta do captcha
def obter_resposta_captcha(api_key, captcha_id):
    url = f"http://2captcha.com/res.php?key={api_key}&action=get&id={captcha_id}"
    while True:
        response = requests.get(url)
        if 'CAPCHA_NOT_READY' in response.text:
            print("Captcha ainda não resolvido, tentando novamente em 5 segundos...")
            time.sleep(5)
        elif 'OK|' in response.text:
            return response.text.split('|')[1]
        else:
            raise Exception(f"Erro ao obter resposta do captcha: {response.text}")

# 3. Inserir o token na página (use Selenium ou outro método)
def inserir_token(token):
    
    # Espera o carregamento da página e insere o token
    navegador.execute_script(f"document.getElementById('g-recaptcha-response').innerHTML = '{token}'")

navegador.get("https://www.transito.mg.gov.br/veiculos/situacao-do-veiculo/consultar-situacao-do-veiculo/")

#Função para contornar erro

def erroSite():
    try:               
        #Valida a tela de erro
        telaErro1 = WebDriverWait(navegador,5).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "body > main > div > div.card > div > div > div > div > h2"))
        )
        navegador.refresh()
                                    
        time.sleep(1)
        pyautogui.hotkey('enter')
                                    
    except TimeoutException:
        print("Prosseguir cadastro")

try:
    
    #ESPERA A TELA PARA REALZIAR O LOGIN NO GOV
    telaParaLoginGov = WebDriverWait(navegador, 15).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "body > main > div > div.row"))
    )
except TimeoutException:
    
    #SE DER ERRO NAVEGA PARA A PAGINA NOVAMENTE
    navegador.get("https://www.transito.mg.gov.br/veiculos/situacao-do-veiculo/consultar-situacao-do-veiculo/")
    
#CLICA NO BOTÃO PARA REALIZAR O LOGIN COM O GOV
botaoLoginGov = navegador.find_element(By.CSS_SELECTOR, "#content > div > div > div.row.text-center.h5.mt-2 > div > a")
navegador.execute_script("arguments[0].click();", botaoLoginGov)

#Espera a primeira tela do login do GOV
tela1GOV = WebDriverWait(navegador, 100).until(
    EC.visibility_of_element_located((By.CSS_SELECTOR, "body > div.container"))
)

# Cria uma instância de ActionChains
actions = ActionChains(navegador)

cpfLogin = "62127152620"
senhaLogin = "covjic-byfpAv-0rapgo"

#Inseri o CPF no campo do CPF
campoCPF = navegador.find_element(By.CSS_SELECTOR, "#accountId")
for caractere in cpfLogin:
    campoCPF.send_keys(caractere)
    time.sleep(0.2)
    

#Clica no botão para prosseguir para outra pagina para colocar a senha
botaoContinuarLogin = navegador.find_element(By.CSS_SELECTOR, "#enter-account-id")
navegador.execute_script("arguments[0].click();", botaoContinuarLogin)

#Espera a segunda tela de login do GOV
tela2GOV = WebDriverWait(navegador, 1000).until(
    EC.visibility_of_element_located((By.CSS_SELECTOR, "#login-password-info > img"))
)


#Inseri a senha no site
campoSENHA = navegador.find_element(By.CSS_SELECTOR, "#password")
for caractere in senhaLogin:
    campoSENHA.send_keys(caractere)
    time.sleep(0.2)


#Entra no site
botaoENTRAR = navegador.find_element(By.CSS_SELECTOR, "#submit-button")
navegador.execute_script("arguments[0].click();", botaoENTRAR)

try:
    
    #Espera a pagina pricipal do DETRAN MG
    paginaPrincipal = WebDriverWait(navegador, 30).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "body div.col-sm.col.px-0"))
    )
    
    #Carrega os cabeçalhos da planilha
    guia_relacao_veiculos['A1'] = "Status"
    guia_relacao_veiculos['B1'] = "Placa"
    guia_relacao_veiculos['C1'] = "Chassi"
    guia_relacao_veiculos['D1'] = "RENAVAM"
    guia_relacao_veiculos['E1'] = "Situação"
    guia_relacao_veiculos['F1'] = "Qtde Autuação"
    guia_relacao_veiculos['G1'] = "Qtde Multas"
    guia_relacao_veiculos['H1'] = "OBS"
    guia_relacao_veiculos['I1'] = "conf. Autuação"
    guia_relacao_veiculos['J1'] = "conf. Multas"
    
    #Passa os valores para uma lista
    status = []
    placa = []
    chassi = []
    renavam = []
    
    time.sleep(2)
    

    linhaResult = 1
        
    for index, row in enumerate(guia_relacao_veiculos.iter_rows(min_row=2, max_row=guia_relacao_veiculos.max_row), start=0):
        linhaResult +=1
        status.append(row[0].value) #Obtem o status
        placa.append(row[1].value) #Obtem o valor da placa
        chassi.append(row[2].value) #Obtem o chassi
        renavam.append(row[3].value) #Obtem renavam
        
        if status[index] is None:
            
            #Inseri a placa no campo placa
            campoPlaca = navegador.find_element(By.CSS_SELECTOR, "#placa")
            campoPlaca.send_keys(placa[index])
            
            #Inseri o chassi no campo chassi
            campoChassi = navegador.find_element(By.CSS_SELECTOR, "#chassi")
            campoChassi.send_keys(chassi[index])
            
            # Função principal
            def resolver_recaptcha():
                try:
                    print("Enviando requisição para resolver reCAPTCHA...")
                    captcha_id = enviar_requisicao_captcha(API_KEY, GOOGLE_KEY, PAGE_URL)
                    
                    print("Aguardando solução do captcha...")
                    token = obter_resposta_captcha(API_KEY, captcha_id)
                    
                    print("Inserindo token e enviando formulário...")
                    inserir_token(token)
                    
                    print("reCAPTCHA resolvido com sucesso!")
                    
                    botaoPEsquisar = navegador.find_element(By.CSS_SELECTOR, "#content > form > button")
                    navegador.execute_script("arguments[0].click();", botaoPEsquisar)
                
                    #=============================================================================================
                    
                    #Passa pelo erro do site
                    erroSite()
                    
                    #Verifica se o veículo não esta cadastrado na base Minas
                    try:
                        
                        campoVeiculoNaoCadastrado = navegador.find_element(By.CSS_SELECTOR, "div.alert.alert-danger.px-3.py-2.font-weight-bold.h5")
                        valorVeiculoNaoCadastrado = campoVeiculoNaoCadastrado.text
                        
                        if (valorVeiculoNaoCadastrado == 'Veículo não Cadastrado na Base de Minas Gerais'):
                            guia_relacao_veiculos[f'H{linhaResult}'] = "Veículo não Cadastrado na Base de Minas Gerais"  
                            planilha.save(caminhoExcel)
                            
                            #Retorna para buscar o proximo veiculo
                            campoConsultaOutroVeiculo = WebDriverWait(navegador, 15).until(
                                EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.text-secondary[href*="/veiculos/situacao-do-veiculo/consultar-situacao-do-veiculo/"]'))
                            ).click()
                            
                            #Coloca o status na planilha
                            guia_relacao_veiculos[f'A{linhaResult}'] = "OK!"  # Inseri o status na planilha
                            planilha.save(caminhoExcel)
                                    
                            return  # Retorna da função para continuar o loop principal

                    except NoSuchElementException:
                        
                        print("Veículo cadastrado na base Minas")
                        
                    time.sleep(3)
                    
                    #                               EMISSÃO DE NOTIFICAÇÃO
                    
                    try:
                        
                        #Botão de autuações
                        campoAutuacao = WebDriverWait(navegador, 15).until(
                            EC.visibility_of_element_located((By.CSS_SELECTOR, "#content > h2:nth-child(6)"))
                        )
                        campoAutuacao.click()
                        
                        #Esperar o modal da autuação abrir 
                        modalAutuacao = WebDriverWait(navegador, 15).until(
                            EC.visibility_of_element_located((By.CSS_SELECTOR, "#divDefesas"))
                        )
                        
                        time.sleep(1)
                        
                        # Localizando a tabela pela classe
                        tabelaAtuacao = navegador.find_element(By.CSS_SELECTOR, ".table.table-sm.table-striped.table-bordered")
                        
                        # Contando o número de linhas (tr)
                        linhasAutuacao = tabelaAtuacao.find_elements(By.TAG_NAME, "tr")
                        numero_linhas = len(linhasAutuacao)
                        
                        quantidadeAutuacoes = 0  # Inicializa a variável para acumular as autuações
                        
                        time.sleep(2)
                        
                        
                        #EMITINDO AS NOTIFICAÕES
                        for cont in range(2, numero_linhas + 1):
                            
                            #Seleciona a quantidade de multas
                            seletorQuantidadeAutuacoes = f"#divDefesas > div > table > tbody > tr:nth-child({cont}) > td.text-center"
                            campoQuantidadeAutuacoes = navegador.find_element(By.CSS_SELECTOR, seletorQuantidadeAutuacoes)
                            
                            # Captura o texto da célula
                            valor = campoQuantidadeAutuacoes.text
                            
                            valor = int(valor)
                            
                            quantidadeAutuacoes += valor
                            
                            time.sleep(2)

                            if (valor == 1):
                                
                                #Clica na multa atual para realizar a emissão
                                seletorEmissaoAutuacao = f"#divDefesas > div > table > tbody > tr:nth-child({cont}) > td > a"
                                campoEmissaoAutuacao = navegador.find_element(By.CSS_SELECTOR, seletorEmissaoAutuacao)
                                navegador.execute_script("arguments[0].click();", campoEmissaoAutuacao)
                            
                                #Pula o erro do site                                erroSite()
                                
                                esperaAutuacao  = WebDriverWait(navegador, 10).until(
                                    EC.visibility_of_element_located((By.CSS_SELECTOR, "#content > dl > dd:nth-child(2)"))
                                )
                                
                                #Obtem o orgão autuador
                                campoOrgaoAutuador = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(2)")
                                valorOrgaoAutuador = campoOrgaoAutuador.text
                                
                                #Obtem a sistuação
                                campoSituacao = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(4)")
                                valorSituacao = campoSituacao.text
                                
                                #Obtem o codigo da multa
                                campoCodigoMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(8)")
                                valorCodigoMulta = campoCodigoMulta.text
                                
                                #Obtem a data da multa
                                campoDataMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(10)")
                                valorDataMulta = campoDataMulta.text
                                
                                #Obtem a hora da multa
                                campoHoraMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(12)")
                                valorHoraMulta = campoHoraMulta.text
                                
                                #Obtem a descrição da multa
                                campoDescricaoMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(14)") 
                                valorDescricaoMulta = campoDescricaoMulta.text
                                
                                #obtem o local da multa 
                                campoLocalMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(16)")
                                valorLocalMulta = campoLocalMulta.text
                                
                                #Obtem o municipio 
                                campoMunicipioMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(18)")
                                valorMunicipioMulta = campoMunicipioMulta.text
                                
                                #Obtem a incluida em
                                campoIncluidaMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(20)")
                                valorIncluidaMulta = campoIncluidaMulta.text
                                                        
                                #Obtem a data limite defesa
                                campoDataLimiteDefesa = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(22)")
                                valorDataLimiteDefesa = campoDataLimiteDefesa.text
                                
                                #Obtem o numero do AIT
                                campoNumeroAIT = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(24)")
                                valorNumeroAIT = campoNumeroAIT.text
                                
                                #Obtem o numero do processamento
                                campoNumeroProcessamento = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(26)")
                                valorNumeroProcessamento = campoNumeroProcessamento.text
                                
                                # Converte os valores da coluna B em um conjunto para facilitar a busca
                                colunaProcessamento = set(cell.value for cell in guia_resultado_autuacao['F'] if cell.value is not None)

                                # Valor que você quer verificar
                                valorProcessamento = valorNumeroProcessamento

                                # Define `existe` como True se o valor estiver no conjunto, caso contrário, False
                                existe = valorProcessamento in colunaProcessamento
            
                                if not existe:
                                    ultimaLinha1 = guia_resultado_autuacao.max_row + 1
                                        
                                    # INSERINDO OS VALORES OBTIDOS DAS AUTUAÇÕES NA PLANILHA
                                    guia_resultado_autuacao[f'A{ultimaLinha1}'] = placa[index]  # Inserindo o valor da placa
                                    guia_resultado_autuacao[f'B{ultimaLinha1}'] = valorNumeroAIT  # Inserindo o número do AIT
                                    guia_resultado_autuacao[f'C{ultimaLinha1}'] = valorDataMulta  # Inserindo o valor da data da multa
                                    guia_resultado_autuacao[f'D{ultimaLinha1}'] = valorHoraMulta  # Inserindo a hora da multa
                                    guia_resultado_autuacao[f'E{ultimaLinha1}'] = valorCodigoMulta  # Inserindo o valor do código da multa
                                    guia_resultado_autuacao[f'F{ultimaLinha1}'] = valorNumeroProcessamento  # Inserindo o número de processamento
                                    guia_resultado_autuacao[f'G{ultimaLinha1}'] = valorSituacao  # Inserindo o valor da situação
                                    guia_resultado_autuacao[f'H{ultimaLinha1}'] = valorDescricaoMulta  # Inserindo o valor da descrição da multa
                                    guia_resultado_autuacao[f'I{ultimaLinha1}'] = valorLocalMulta  # Inserindo o valor do local da multa
                                    guia_resultado_autuacao[f'J{ultimaLinha1}'] = valorMunicipioMulta  # Inserindo o valor do município na multa
                                    guia_resultado_autuacao[f'K{ultimaLinha1}'] = valorIncluidaMulta  # Inserindo o valor de quando a multa é incluída
                                    guia_resultado_autuacao[f'L{ultimaLinha1}'] = valorDataLimiteDefesa  # Inserindo o valor da data limite da multa
                                    guia_resultado_autuacao[f'M{ultimaLinha1}'] = valorOrgaoAutuador  # Inserindo o valor do órgão autuador
                                    
                                    #Salva os resultados na planilha
                                    planilha.save(caminhoExcel)
                                    
                                    time.sleep(2)
                                
                                #Clica para consulta a proxima autuação
                                campoVoltarDadosVeiculos = navegador.find_element(By.CSS_SELECTOR, "#content > div:nth-child(8) > a")
                                navegador.execute_script("arguments[0].click();", campoVoltarDadosVeiculos)
                                
                                time.sleep(1)
                                
                                #Valida a tela de erro
                                erroSite()
                                
                                #Botão de autuações
                                campoAutuacao = WebDriverWait(navegador, 15).until(
                                        EC.visibility_of_element_located((By.CSS_SELECTOR, "#content > h2:nth-child(6)"))
                                )
                                campoAutuacao.click()
                                
                                time.sleep(2)
                                
                            else:
                                
                                time.sleep(2)
                                
                                #Clica na multa atual para realizar a emissão
                                seletorEmissaoAutuacao = f"#divDefesas > div > table > tbody > tr:nth-child({cont}) > td > a"
                                campoEmissaoAutuacao = navegador.find_element(By.CSS_SELECTOR, seletorEmissaoAutuacao)
                                navegador.execute_script("arguments[0].click();", campoEmissaoAutuacao)

                                #Valida a tela de erro
                                erroSite()
                                
                                esperaAutuacao  = WebDriverWait(navegador, 10).until(
                                    EC.visibility_of_element_located((By.CSS_SELECTOR, "#content > dl > dd:nth-child(2)"))
                                )
                                
                                for i in range(0, valor):
                                    
                                    #Valida a tela de erro
                                    erroSite()    
                                    
                                    #Obtem o orgão autuador
                                    campoOrgaoAutuador = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(2)")
                                    valorOrgaoAutuador = campoOrgaoAutuador.text
                                    
                                    #Obtem a sistuação
                                    campoSituacao = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(4)")
                                    valorSituacao = campoSituacao.text
                                    
                                    #Obtem o codigo da multa
                                    campoCodigoMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(8)")
                                    valorCodigoMulta = campoCodigoMulta.text
                                    
                                    #Obtem a data da multa
                                    campoDataMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(10)")
                                    valorDataMulta = campoDataMulta.text
                                    
                                    #Obtem a hora da multa
                                    campoHoraMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(12)")
                                    valorHoraMulta = campoHoraMulta.text
                                    
                                    #Obtem a descrição da multa
                                    campoDescricaoMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(14)") 
                                    valorDescricaoMulta = campoDescricaoMulta.text
                                    
                                    #obtem o local da multa 
                                    campoLocalMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(16)")
                                    valorLocalMulta = campoLocalMulta.text
                                    
                                    #Obtem o municipio 
                                    campoMunicipioMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(18)")
                                    valorMunicipioMulta = campoMunicipioMulta.text
                                    
                                    #Obtem a incluida em
                                    campoIncluidaMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(20)")
                                    valorIncluidaMulta = campoIncluidaMulta.text
                                                            
                                    #Obtem a data limite defesa
                                    campoDataLimiteDefesa = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(22)")
                                    valorDataLimiteDefesa = campoDataLimiteDefesa.text
                                    
                                    #Obtem o numero do AIT
                                    campoNumeroAIT = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(24)")
                                    valorNumeroAIT = campoNumeroAIT.text
                                    
                                    #Obtem o numero do processamento
                                    campoNumeroProcessamento = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(26)")
                                    valorNumeroProcessamento = campoNumeroProcessamento.text
                                    
                                    # Converte os valores da coluna B em um conjunto para facilitar a busca
                                    colunaProcessamento = set(cell.value for cell in guia_resultado_autuacao['F'] if cell.value is not None)

                                    # Valor que você quer verificar
                                    valorProcessamento = valorNumeroProcessamento

                                    # Define `existe` como True se o valor estiver no conjunto, caso contrário, False
                                    existe = valorProcessamento in colunaProcessamento
            
                                    if not existe:
                                          
                                        ultimaLinha2 = guia_resultado_autuacao.max_row + 1
                                            
                                        # INSERINDO OS VALORES OBTIDOS DAS AUTUAÇÕES NA PLANILHA
                                        guia_resultado_autuacao[f'A{ultimaLinha2}'] = placa[index]  # Inserindo o valor da placa
                                        guia_resultado_autuacao[f'B{ultimaLinha2}'] = valorNumeroAIT  # Inserindo o número do AIT
                                        guia_resultado_autuacao[f'C{ultimaLinha2}'] = valorDataMulta  # Inserindo o valor da data da multa
                                        guia_resultado_autuacao[f'D{ultimaLinha2}'] = valorHoraMulta  # Inserindo a hora da multa
                                        guia_resultado_autuacao[f'E{ultimaLinha2}'] = valorCodigoMulta  # Inserindo o valor do código da multa
                                        guia_resultado_autuacao[f'F{ultimaLinha2}'] = valorNumeroProcessamento  # Inserindo o número de processamento
                                        guia_resultado_autuacao[f'G{ultimaLinha2}'] = valorSituacao  # Inserindo o valor da situação
                                        guia_resultado_autuacao[f'H{ultimaLinha2}'] = valorDescricaoMulta  # Inserindo o valor da descrição da multa
                                        guia_resultado_autuacao[f'I{ultimaLinha2}'] = valorLocalMulta  # Inserindo o valor do local da multa
                                        guia_resultado_autuacao[f'J{ultimaLinha2}'] = valorMunicipioMulta  # Inserindo o valor do município na multa
                                        guia_resultado_autuacao[f'K{ultimaLinha2}'] = valorIncluidaMulta  # Inserindo o valor de quando a multa é incluída
                                        guia_resultado_autuacao[f'L{ultimaLinha2}'] = valorDataLimiteDefesa  # Inserindo o valor da data limite da multa
                                        guia_resultado_autuacao[f'M{ultimaLinha2}'] = valorOrgaoAutuador  # Inserindo o valor do órgão autuador
                                        
                                        #Salva os resultados na planilha
                                        planilha.save(caminhoExcel)
                                        
                                        time.sleep(3)
                                    
                                    try:
                                        
                                        #Clica no botão proxima autuação
                                        botaoProxAutuacao = navegador.find_element(By.XPATH, '//*[@id="content"]/div[3]/div/a')
                                        navegador.execute_script("arguments[0].click();", botaoProxAutuacao)
                                    except:
                                        continue
                                    
                                #Clica para consulta a proxima autuação
                                campoVoltarDadosVeiculos = navegador.find_element(By.CSS_SELECTOR, "#content > div:nth-child(8) > a")
                                navegador.execute_script("arguments[0].click();", campoVoltarDadosVeiculos)
                                
                                time.sleep(2)
                                
                                #Valida a tela de erro
                                erroSite()
                                
                                
                                #Botão de autuações
                                campoAutuacao = WebDriverWait(navegador, 15).until(
                                        EC.visibility_of_element_located((By.CSS_SELECTOR, "#content > h2:nth-child(6)"))
                                )
                                campoAutuacao.click()
                                
                                time.sleep(2)

                        time.sleep(2)
                        guia_relacao_veiculos[f'F{linhaResult}'] = quantidadeAutuacoes  # Inserindo a quantidade de autuações
                        planilha.save(caminhoExcel)
                    
                    except TimeoutException:
                        guia_relacao_veiculos[f'F{linhaResult}'] = 0  # Inserindo a quantidade de autuações
                        planilha.save(caminhoExcel)
                        print("Veículo sem notificação")
                        
                    #limpa a variavel numero_linhas
                    numero_linhas = 0
                    valor = 0
                    
                    
                    #                               EMISSÃO DE MULTAS
                    
                    try:
                        
                        #Espera a janela de multas
                        campoMultas = WebDriverWait(navegador, 15).until(
                            EC.visibility_of_element_located((By.CSS_SELECTOR, "#content > h2:nth-child(8)"))
                        ).click()
                        
                        #Esperar o modal da autuação abrir 
                        modalMultas = WebDriverWait(navegador, 15).until(
                            EC.visibility_of_element_located((By.CSS_SELECTOR, "#divMultas"))
                        )
                        
                        time.sleep(1)
                        
                        tabelaMultas = WebDriverWait(navegador, 10).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "#divMultas > div > table"))
                        )
                        
                        
                        # Contando o número de linhas (tr)
                        linhasMultas = tabelaMultas.find_elements(By.TAG_NAME, "tr")
                        numero_linhas = len(linhasMultas)
                        

                        # Inicializa a variável para acumular as autuações
                        quantidadeMultas = 0  
                        
                        time.sleep(1.5)
                        
                        for cont in range(2, numero_linhas + 1):  
                            #Seleciona a quantidade de multas
                            seletorQuantidadeMultas = f"#divMultas > div > table > tbody > tr:nth-child({cont}) > td.text-center"
                            campoQuantidadeMultas = navegador.find_element(By.CSS_SELECTOR, seletorQuantidadeMultas)
                            
                            # Captura o texto da célula
                            valor = campoQuantidadeMultas.text
                            
                            valor = int(valor)
                            
                            quantidadeMultas += valor
                            
                            time.sleep(2)
                            
                            if (valor == 1):
                                
                                #Clica na multa atual para realizar a emissão
                                seletorEmissaoMulta = f"#divMultas > div > table > tbody > tr:nth-child({cont}) > td > a"
                                campoEmissaoMulta = navegador.find_element(By.CSS_SELECTOR, seletorEmissaoMulta)
                                navegador.execute_script("arguments[0].click();", campoEmissaoMulta)
                            
                                #Valida a tela de erro
                                erroSite()
                                
                                esperaMulta  = WebDriverWait(navegador, 10).until(
                                    EC.visibility_of_element_located((By.CSS_SELECTOR, "#content > dl > dd:nth-child(2)"))
                                )
                                
                                #Obtem o orgão autuador
                                campoOrgaoAutuador = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(1) > dd")
                                valorOrgaoAutuador = campoOrgaoAutuador.text
                                    
                                #Obtem a sistuação
                                campoSituacao = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(2) > dd")
                                valorSituacao = campoSituacao.text
                                    
                                #Obtem o codigo da multa
                                campoCodigoMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(4) > dd")
                                valorCodigoMulta = campoCodigoMulta.text
                                    
                                #Obtem a data da multa
                                campoDataMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(5) > dd")
                                valorDataMulta = campoDataMulta.text
                                    
                                #Obtem a hora da multa
                                campoHoraMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(6) > dd")
                                valorHoraMulta = campoHoraMulta.text
                                    
                                #Obtem a descrição da multa
                                campoDescricaoMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(7) > dd") 
                                valorDescricaoMulta = campoDescricaoMulta.text
                                    
                                #obtem o local da multa 
                                campoLocalMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(8) > dd")
                                valorLocalMulta = campoLocalMulta.text
                                    
                                #Obtem o municipio 
                                campoMunicipioMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(9) > dd")
                                valorMunicipioMulta = campoMunicipioMulta.text
                                    
                                #Obtem a incluida em
                                campoIncluidaMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(10) > dd")
                                valorIncluidaMulta = campoIncluidaMulta.text
                                                            
                                #Obtem a data limite defesa
                                campoDataLimiteDefesa = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(11) > dd")
                                valorDataLimiteDefesa = campoDataLimiteDefesa.text
                                    
                                #Obtem o numero do AIT
                                campoNumeroAIT = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(12) > dd")
                                valorNumeroAIT = campoNumeroAIT.text
                                    
                                #Obtem o numero do processamento
                                campoNumeroProcessamento = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(13) > dd")
                                valorNumeroProcessamento = campoNumeroProcessamento.text
                                
                                # Converte os valores da coluna B em um conjunto para facilitar a busca
                                colunaProcessamento = set(cell.value for cell in guia_resultado_autuacao['F'] if cell.value is not None)

                                # Valor que você quer verificar
                                valorProcessamento = valorNumeroProcessamento

                                # Define `existe` como True se o valor estiver no conjunto, caso contrário, False
                                existe = valorProcessamento in colunaProcessamento
            
                                if not existe:
                                        
                                    ultimaLinha3 = guia_resultado_multa.max_row + 1
                                    
                                    # INSERINDO OS VALORES OBTIDOS DAS MULTAS NA PLANILHA
                                    guia_resultado_multa[f'A{ultimaLinha3}'] = placa[index]  # Inserindo o valor da placa
                                    guia_resultado_multa[f'B{ultimaLinha3}'] = valorNumeroAIT  # Inserindo o número do AIT
                                    guia_resultado_multa[f'C{ultimaLinha3}'] = valorDataMulta  # Inserindo o valor da data da multa
                                    guia_resultado_multa[f'D{ultimaLinha3}'] = valorHoraMulta  # Inserindo a hora da multa
                                    guia_resultado_multa[f'E{ultimaLinha3}'] = valorCodigoMulta  # Inserindo o valor do código da multa
                                    guia_resultado_multa[f'F{ultimaLinha3}'] = valorNumeroProcessamento  # Inserindo o número de processamento
                                    guia_resultado_multa[f'G{ultimaLinha3}'] = valorSituacao  # Inserindo o valor da situação
                                    guia_resultado_multa[f'H{ultimaLinha3}'] = valorDescricaoMulta  # Inserindo o valor da descrição da multa
                                    guia_resultado_multa[f'I{ultimaLinha3}'] = valorLocalMulta  # Inserindo o valor do local da multa
                                    guia_resultado_multa[f'J{ultimaLinha3}'] = valorMunicipioMulta  # Inserindo o valor do município na multa
                                    guia_resultado_multa[f'K{ultimaLinha3}'] = valorIncluidaMulta  # Inserindo o valor de quando a multa é incluída
                                    guia_resultado_multa[f'L{ultimaLinha3}'] = valorDataLimiteDefesa  # Inserindo o valor da data limite da multa
                                    guia_resultado_multa[f'M{ultimaLinha3}'] = valorOrgaoAutuador  # Inserindo o valor do órgão autuador
                                    guia_resultado_multa[f'N{ultimaLinha3}'] = valorMulta  # Inserindo o valor da multa-
                                    
                                    #Salva os resultados na planilha
                                    planilha.save(caminhoExcel)
                                    
                                    time.sleep(2)
                            
                                #Clica para consulta a proxima multa
                                campoVoltarDadosVeiculos = navegador.find_element(By.CSS_SELECTOR, "#content > div:nth-child(8) > a")
                                navegador.execute_script("arguments[0].click();", campoVoltarDadosVeiculos)
                                
                                time.sleep(1)
                                
                                #Valida a tela de erro
                                erroSite()
                                            
                                #Botão de autuações
                                campoMulta = WebDriverWait(navegador, 15).until(
                                        EC.visibility_of_element_located((By.CSS_SELECTOR, "#content > h2:nth-child(8)"))
                                )
                                campoMulta.click()
                                
                                time.sleep(2)
                                
                            else:
                                
                                time.sleep(2)
                                #Clica na multa atual para realizar a emissão
                                seletorEmissaoMulta = f"#divMultas > div > table > tbody > tr:nth-child({cont}) > td > a"
                                campoEmissaoMulta = navegador.find_element(By.CSS_SELECTOR, seletorEmissaoMulta)
                                navegador.execute_script("arguments[0].click();", campoEmissaoMulta)

                                #Valida a tela de erro
                                erroSite()    
                                    
                                esperaMulta  = WebDriverWait(navegador, 10).until(
                                    EC.visibility_of_element_located((By.CSS_SELECTOR, "#content  dl:nth-child(1) > dt"))
                                )
                                
                                for i in range(0, valor):
                                    
                                    #Valida a tela de erro
                                    erroSite()    
                                    
                                    #Obtem o orgão autuador
                                    campoOrgaoAutuador = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(1) > dd")
                                    valorOrgaoAutuador = campoOrgaoAutuador.text
                                    
                                    #Obtem a sistuação
                                    campoSituacao = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(2) > dd")
                                    valorSituacao = campoSituacao.text
                                    
                                    #Obtem o codigo da multa
                                    campoCodigoMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(4) > dd")
                                    valorCodigoMulta = campoCodigoMulta.text
                                    
                                    #Obtem a data da multa
                                    campoDataMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(5) > dd")
                                    valorDataMulta = campoDataMulta.text
                                    
                                    #Obtem a hora da multa
                                    campoHoraMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(6) > dd")
                                    valorHoraMulta = campoHoraMulta.text
                                    
                                    #Obtem a descrição da multa
                                    campoDescricaoMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(7) > dd") 
                                    valorDescricaoMulta = campoDescricaoMulta.text
                                    
                                    #obtem o local da multa 
                                    campoLocalMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(8) > dd")
                                    valorLocalMulta = campoLocalMulta.text
                                    
                                    #Obtem o municipio 
                                    campoMunicipioMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(9) > dd")
                                    valorMunicipioMulta = campoMunicipioMulta.text
                                    
                                    #Obtem a incluida em
                                    campoIncluidaMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(10) > dd")
                                    valorIncluidaMulta = campoIncluidaMulta.text
                                                            
                                    #Obtem a data limite defesa
                                    campoDataLimiteDefesa = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(11) > dd")
                                    valorDataLimiteDefesa = campoDataLimiteDefesa.text
                                    
                                    #Obtem o numero do AIT
                                    campoNumeroAIT = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(12) > dd")
                                    valorNumeroAIT = campoNumeroAIT.text
                                    
                                    #Obtem o numero do processamento
                                    campoNumeroProcessamento = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(13) > dd")
                                    valorNumeroProcessamento = campoNumeroProcessamento.text
                                    
                                    #Obtem o valor da multa
                                    campoValorMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(14) > dd")
                                    valorMulta = campoValorMulta.text
                                        
                                    # Converte os valores da coluna B em um conjunto para facilitar a busca
                                    colunaProcessamento = set(cell.value for cell in guia_resultado_autuacao['F'] if cell.value is not None)

                                    # Valor que você quer verificar
                                    valorProcessamento = valorNumeroProcessamento

                                    # Define `existe` como True se o valor estiver no conjunto, caso contrário, False
                                    existe = valorProcessamento in colunaProcessamento
                
                                    if not existe:
                                        
                                        ultimaLinha4 = guia_resultado_multa.max_row + 1
                                        
                                        # INSERINDO OS VALORES OBTIDOS DAS MULTAS NA PLANILHA
                                        guia_resultado_multa[f'A{ultimaLinha4}'] = placa[index]  # Inserindo o valor da placa
                                        guia_resultado_multa[f'B{ultimaLinha4}'] = valorNumeroAIT  # Inserindo o número do AIT
                                        guia_resultado_multa[f'C{ultimaLinha4}'] = valorDataMulta  # Inserindo o valor da data da multa
                                        guia_resultado_multa[f'D{ultimaLinha4}'] = valorHoraMulta  # Inserindo a hora da multa
                                        guia_resultado_multa[f'E{ultimaLinha4}'] = valorCodigoMulta  # Inserindo o valor do código da multa
                                        guia_resultado_multa[f'F{ultimaLinha4}'] = valorNumeroProcessamento  # Inserindo o número de processamento
                                        guia_resultado_multa[f'G{ultimaLinha4}'] = valorSituacao  # Inserindo o valor da situação
                                        guia_resultado_multa[f'H{ultimaLinha4}'] = valorDescricaoMulta  # Inserindo o valor da descrição da multa
                                        guia_resultado_multa[f'I{ultimaLinha4}'] = valorLocalMulta  # Inserindo o valor do local da multa
                                        guia_resultado_multa[f'J{ultimaLinha4}'] = valorMunicipioMulta  # Inserindo o valor do município na multa
                                        guia_resultado_multa[f'K{ultimaLinha4}'] = valorIncluidaMulta  # Inserindo o valor de quando a multa é incluída
                                        guia_resultado_multa[f'L{ultimaLinha4}'] = valorDataLimiteDefesa  # Inserindo o valor da data limite da multa
                                        guia_resultado_multa[f'M{ultimaLinha4}'] = valorOrgaoAutuador  # Inserindo o valor do órgão autuador
                                        guia_resultado_multa[f'N{ultimaLinha4}'] = valorMulta  # Inserindo o valor da multa
                                    
                                        #Salva os resultados na planilha
                                        planilha.save(caminhoExcel)
                                        
                                        time.sleep(3)
                                    
                                    try:
                                        
                                        #Clica no botão proxima autuação
                                        botaoProxMulta = navegador.find_element(By.XPATH, '//*[@id="content"]/div[4]/div/a')
                                        navegador.execute_script("arguments[0].click();", botaoProxMulta)
                                    except:
                                            continue
                                    
                                #Clica para consulta a proxima autuação
                                campoVoltarDadosVeiculos = navegador.find_element(By.CSS_SELECTOR, "#content > div:nth-child(8) > a")
                                navegador.execute_script("arguments[0].click();", campoVoltarDadosVeiculos)
                                
                                time.sleep(2)
                                
                                #Valida a tela de erro
                                erroSite()    
                                
                                #Botão de autuações
                                campoMulta = WebDriverWait(navegador, 15).until(
                                        EC.visibility_of_element_located((By.CSS_SELECTOR, "#content > h2:nth-child(8)"))
                                )
                                campoMulta.click()
                                
                                time.sleep(2)
                            
                        time.sleep(2)
                        guia_relacao_veiculos[f'G{linhaResult}'] = quantidadeMultas  # Inserindo a quantidade de autuações
                        planilha.save(caminhoExcel)    
                            
                    except TimeoutException:
                        guia_relacao_veiculos[f'G{linhaResult}'] = 0  # Inserindo a quantidade de autuações
                        planilha.save(caminhoExcel)
                        print("Veículo sem multas")
                        
                    #Retorna para buscar o proximo veiculo
                    campoConsultaOutroVeiculo = WebDriverWait(navegador, 15).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.text-secondary[href*="/veiculos/situacao-do-veiculo/consultar-situacao-do-veiculo/"]'))
                    ).click()
                    
                    #Coloca o status na planilha
                    guia_relacao_veiculos[f'A{linhaResult}'] = "OK!"  # Inseri o status na planilha
                    planilha.save(caminhoExcel)
                    
                    
                except Exception as e:
                    print(f"ERRO AO RESOLVER CAPTCHA {e}")
                    

            # Executar
            resolver_recaptcha()
                
except TimeoutException:
    print("ERRO")
    
navegador.quit()
