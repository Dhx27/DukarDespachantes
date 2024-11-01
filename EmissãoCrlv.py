import os
import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from twocaptcha import TwoCaptcha
import pyautogui
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager  # Importa o WebDriver Manager
from selenium.webdriver.chrome.service import Service  # Importar Service

# Passa o caminho da planilha
caminhoPlanilha = r'C:\Users\diogo.lana\Desktop\TESTE\PLANILHA MODELO (2).xlsx'

# Passa o caminho dos downloads
pathDownloads = r'C:\Users\diogo.lana\Desktop\TESTE'

# region Código Oculto

#o REGION fuciona como uma rotina do ibm
# Configuração da chave da API do 2Captcha
api_key = os.getenv('API_KEY_2CAPTCHA', '657c1d808a967e254d096cd0cfd696c3')
solver = TwoCaptcha(api_key)

# Inicializa o navegador usando o WebDriver Manager
service = Service(ChromeDriverManager().install())  # Cria o serviço do ChromeDriver
navegador = webdriver.Chrome(service=service)  # Passa o serviço na inicialização do webdriver
navegador.maximize_window()

# Acessa o site
navegador.get("https://transito.mg.gov.br/")

# Navega até a página desejada
botaoVeiculos = navegador.find_element(By.CSS_SELECTOR, "#nav > ul > li.nav-item.yamm-fw.menu-icons-veiculos > a")
botaoVeiculos.click()

time.sleep(2)

botaoImprimirCRLV = navegador.find_element(By.CSS_SELECTOR, "body > main > div > div.row > div.col-md-8 > div > div.card-body > div > div:nth-child(6) > ul > li:nth-child(2) > a > span")
botaoImprimirCRLV.click()

# Carrega a planilha
workbook = load_workbook(caminhoPlanilha)
sheet = workbook.active

# Itera sobre as linhas da planilha
for linhaDaVez in range(2, sheet.max_row + 1):  # Assume que a primeira linha é o cabeçalho
    placa = sheet[f'A{linhaDaVez}'].value  # Supondo que 'PLACA' está na coluna A
    renavam = sheet[f'B{linhaDaVez}'].value  # Supondo que 'RENAVAM' está na coluna B
    cnpj = sheet[f'C{linhaDaVez}'].value  # Supondo que 'CNPJ' está na coluna C
    numeroCrv = sheet[f'D{linhaDaVez}'].value  # Supondo que 'NÚMERO DE CRV' está na coluna D
    status = sheet[f'E{linhaDaVez}'].value  # Supondo que 'STATUS' está na coluna E

    if status != "OK!":
        
        print("Pesquisa placa: " + placa)

        campoPlaca = navegador.find_element(By.ID, "placa")
        campoPlaca.send_keys(placa)
    
        campoRenavam = navegador.find_element(By.ID, "renavam")
        campoRenavam.send_keys(renavam)
        
        campoCNPJ = navegador.find_element(By.ID, "cpf-cnpj")
        campoCNPJ.send_keys(cnpj)
        
        campoCRV = navegador.find_element(By.ID, "numero-crv")
        campoCRV.send_keys(numeroCrv)

        try:
            # Tentar resolver o reCAPTCHA
            result = solver.recaptcha(
                sitekey='6LfVpnIUAAAAAHkISk6Z6juZcsUx6hbyJGwfnfPL',
                url='https://transito.mg.gov.br/veiculos/documentos-de-veiculos/imprimir-crlv'
            )
            
            # Obtém o token resolvido
            token = result['code']
            print('reCAPTCHA quebrado!')

            # Injeta o token do reCAPTCHA no campo invisível
            navegador.execute_script(f"document.getElementById('g-recaptcha-response').innerHTML = '{token}'")

            # Forçar clique no botão usando JavaScript
            botaoPesquisa = navegador.find_element(By.CSS_SELECTOR, "#form-emitir_crlve > button")
            navegador.execute_script("arguments[0].click();", botaoPesquisa)

            # Espera a tela do boleto aparecer certinho, esperar ele estar visível na tela
            telaBoleto = WebDriverWait(navegador, 60).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, "body > embed"))
            )
            
            time.sleep(5)
            
            # Tenta usar o PyAutoGUI para enviar o atalho Ctrl + P
            pyautogui.hotkey('ctrl', 'p')
            
            time.sleep(2)
            
            pyautogui.hotkey('enter')
            
            time.sleep(2)
        
            pyautogui.write(pathDownloads + '\\' + placa)
            
            pyautogui.hotkey('enter')
            
            time.sleep(3)  # Aguarda um pouco para visualizar a impressão
            
            sheet[f'E{linhaDaVez}'] = 'OK!'  # Atualiza o status na coluna E
            workbook.save(caminhoPlanilha)  # Salva as alterações na planilha
            
            navegador.get("https://transito.mg.gov.br/veiculos/documentos-de-veiculos/imprimir-crlv")
            
            time.sleep(2)  # Aguarda um pouco para visualizar a impressão
            
        except Exception as e:       
            print(f"Erro ao resolver o reCAPTCHA ou imprimir: {e}")
            time.sleep(5)  # Tempo de espera antes de tentar resolver novamente

# Encerra o navegador após completar o processo
navegador.quit()

# endregion
