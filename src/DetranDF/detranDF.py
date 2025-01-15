#ROBO DE EMISSÃO DE IPVA E LICENCIAMENTO Diogo
from dotenv import load_dotenv
from openpyxl import load_workbook
import time
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import undetected_chromedriver as uc
import requests
from selenium.webdriver.common.keys import Keys  # Para acessar as teclas especiais
from selenium.webdriver.common.action_chains import ActionChains
import os
import pyautogui
import pdfplumber
import re
from PyPDF2 import PdfReader, PdfWriter

load_dotenv()

pasta_download = r'C:\Users\Robo01\Desktop\ENTRADA\SAIDA'
caminho_planilha = r'C:\Users\Robo01\Desktop\ENTRADA\MODELO DF.xlsx'

def criar_pasta_saida(pasta_download, caminho_planilha):
    
    # Pega o nome do arquivo da planilha sem a extensão
    nome_arquivo = os.path.splitext(os.path.basename(caminho_planilha))[0]

    #Caminho da pasta de saída com base na pasta fornecida
    pasta_saida = os.path.join(pasta_download, nome_arquivo)

    if not os.path.exists(pasta_saida):

        os.makedirs(pasta_saida)
        print(f"Pasta criada: {pasta_saida}")
    else:
        print(f"a pasta já existe: {pasta_saida}")

    return pasta_saida

# Caminha para selecionar o download como PDF
def selecionar_download_como_pdf(pasta_saida, placa_atual, ano_ipva):

    # Navegar pelas opções até "Salvar"
    for _ in range(5):  # Pressiona 'tab' 5 vezes
        pyautogui.hotkey('tab')
    time.sleep(1.5)

    pyautogui.write("salvar")  # Digita "salvar"
    time.sleep(1.5)

        # Navegar até o botão de salvar
    for _ in range(6):  # Pressiona 'tab' 6 vezes
        pyautogui.hotkey('tab')
    time.sleep(1.5)

    pyautogui.hotkey('enter')  # Confirma "Salvar"
    time.sleep(3)

    # Define o caminho completo para salvar o arquivo
    caminho_download = os.path.join(pasta_saida, f"LIC {placa_atual}_{ano_ipva}")
    caminho_download = os.path.normpath(caminho_download)  # Normaliza o caminho

    # Digitar o caminho de salvamento
    pyautogui.write(caminho_download)
    time.sleep(1)
    pyautogui.hotkey('enter')  # Confirma o caminho

    print("Arquivo salvo com sucesso!")

pasta_saida = criar_pasta_saida(pasta_download, caminho_planilha)

#Abri a planilha do excel 
planilha = load_workbook(caminho_planilha)

#Passa a instacia da planilha BASE
guia_dados = planilha['BASE']


index = 0
linhas = list(guia_dados.iter_rows(min_row=2, max_row=guia_dados.max_row))

# Configurar o Chrome com um User-Agent falso usando undetected-chromedriver
chrome_options = Options()
chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36")

# Inicializa o navegador com as opções e com o stealth mode ativado
navegador = uc.Chrome(options=chrome_options)
navegador.maximize_window()

navegador.get("https://portal.detran.df.gov.br/#/home")

try:

    botao_login = WebDriverWait(navegador, 60).until(
        EC.element_to_be_clickable((By.XPATH, "/html/body/app-root/app-header/mat-toolbar/span/span/span[3]/button"))
    ).click()

    botao_login_gov = WebDriverWait(navegador, 60).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "#enviar"))
    ).click()
    
    campoCPF = WebDriverWait(navegador, 60).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "#accountId"))
    )
    campoCPF.send_keys(os.getenv("CPF_GOIAS"))

    botaoContinuarLogin = navegador.find_element(By.CSS_SELECTOR, "#enter-account-id")
    botaoContinuarLogin.click()

    campoSenha =  WebDriverWait(navegador, 1000).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "#password"))
    )
    campoSenha.send_keys(os.getenv("SENHA_GOIAS"))

    botaoEntrar = navegador.find_element(By.CSS_SELECTOR, "#submit-button")
    botaoEntrar.click()

    try:

        botao_autorizacao = WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "#authorize-info > div.button-panel > button.button-ok"))
        ).click()
    except Exception:
        pass

    time.sleep(5)

    tela_detran_df = WebDriverWait(navegador, 60).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "app-home > section > div:nth-child(1)"))
    )

    time.sleep(2)

    pyautogui.leftClick(x=1188, y=365)

    time.sleep(1)

    cabecalho_veiculos = WebDriverWait(navegador, 60).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "body > app-root > app-header > mat-toolbar > span > span > section > button:nth-child(2)"))
    ).click()

    time.sleep(3)

    botao_consulta_debitos = WebDriverWait(navegador, 60).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "#headerMenu_veiculosConsultas > app-menu-item:nth-child(1) > div.mat-tooltip-trigger.menu-item-wrapper > a > article > h3"))
    ).click()

    botao_consulta_veiculos = WebDriverWait(navegador, 60).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "app-botao-consulta-veiculo-outro-proprietario > div > button"))
    ).click()

    linhaPlan =1

    while index < len(linhas):

        linhaPlan +=1
        row = linhas[index]

        placa_atual = row[0].value
        renavam_atual = row[1].value
        status_atual = row[2].value

        if status_atual is None:

            campo_placa = navegador.find_element(By.XPATH, '(//input[contains(@class, "mat-input-element")])[1]')
            campo_placa.send_keys(placa_atual)

            time.sleep(2)

            campo_renavam = navegador.find_element(By.XPATH, '(//input[contains(@class, "mat-input-element")])[2]')
            campo_renavam.send_keys(renavam_atual)

            time.sleep(1.5)

            botao_consultar = navegador.find_element(By.CSS_SELECTOR, " app-consulta-placa-renavam > mat-card-actions > div > button")
            botao_consultar.click()

            tela_dados_veiculo = WebDriverWait(navegador, 60).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, "mat-card-content > app-lista-dados"))
            )

            botao_consultar_2 = navegador.find_element(By.XPATH, "/html/body/app-root/div/div[1]/app-layout-servicos/app-debitos/app-card-detran/mat-card/mat-card-content/app-lista-dados/div/div[2]/table/tbody/tr/td[7]/button")
            botao_consultar_2.click()

            tela_licenciamento = WebDriverWait(navegador, 60).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, "app-lista-dados:nth-child(2) div.forms-wrapper-header.ng-star-inserted > div > span"))
            )

            time.sleep(3)

            tabela_ipva = navegador.find_element(By.CSS_SELECTOR, "div:nth-child(2) > app-lista-dados:nth-child(2) > div > div.forms-wrapper-body.forms-wrapper-body_table.pad-botton.ng-star-inserted > table")
            linhas_ipva = tabela_ipva.find_elements(By.TAG_NAME, "tr")
            numero_linhas_ipva = len(linhas_ipva)

            
            #                                   EMISSÃO LICENCIAMENTO

            for cont in range(1,  numero_linhas_ipva):

                selector_ano_ipva = f" div.forms-wrapper-body.forms-wrapper-body_table.pad-botton.ng-star-inserted > table > tbody > tr:nth-child({cont}) > td.mat-cell.cdk-cell.cdk-column-Ano.mat-column-Ano.ng-star-inserted"
                selector_botao_emitir = f"div.forms-wrapper-body.forms-wrapper-body_table.pad-botton.ng-star-inserted > table > tbody > tr:nth-child({cont}) > td.mat-cell.cdk-cell.cdk-column-A--o.mat-column-A--o.ng-star-inserted > button"

                campo_ano_ipva = navegador.find_element(By.CSS_SELECTOR, selector_ano_ipva)
                ano_ipva = campo_ano_ipva.text

                botao_emitir = navegador.find_element(By.CSS_SELECTOR, selector_botao_emitir)
                botao_emitir.click()

                tela_boleto = WebDriverWait(navegador, 60).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, "app-invoice > section > div.pagina-bordero"))
                )

                time.sleep(3)
                try:
                    botao_imprimir = navegador.find_element(By.ID, "buttonImprimir")
                    navegador.execute_script("arguments[0].click();", botao_imprimir)
                
                except (NoSuchElementException, TimeoutException): 

                    selecionar_download_como_pdf(pasta_saida, placa_atual, ano_ipva)

                time.sleep(1)

                botao_voltar = WebDriverWait(navegador, 60).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "#footer > mat-card-actions > button"))
                ).click()

            guia_dados[f"C{linhaPlan}"] = "OK!"
            planilha.save(caminho_planilha)
            

            #                                           EMISSÃO IPVA

            botao_ipva_execicio = WebDriverWait(navegador, 60).until(
                EC.element_to_be_clickable((By.XPATH, "/html/body/app-root/div/div[1]/app-layout-servicos/app-debitos/app-card-detran/mat-card/mat-card-content/div[2]/app-ipva/div/div[2]/div[1]/p/mat-form-field/div/div[1]/div[3]/mat-select"))
            ).click()
             
            time.sleep(2)

            parant_element = navegador.find_element(By.ID, 'cdk-overlay-1')

            child_elements = parant_element.find_elements(By.CSS_SELECTOR, '[id^="mat-option-"]')

            print(f"{len(child_elements)}")

            for cont in range(1, len(child_elements) + 1):
                selector_ipva = f"/html/body/div[2]/div[2]/div/div/div/mat-option[{cont}]"

                botao_ano_ipva = navegador.find_element(By.XPATH, selector_ipva)
                botao_ano_ipva.click()

                time.sleep(3)

                selector_ano_ipva = navegador.find_element(By.CSS_SELECTOR, "div:nth-child(3) > app-ipva > div > div.forms-wrapper-body > div.forms-wrapper-body.forms-wrapper-body_table.pad-botton.ng-star-inserted > table > tbody > tr:nth-child(1) > td.mat-cell.cdk-cell.cdk-column-Ano.mat-column-Ano.ng-star-inserted")
                texto_ipva = selector_ano_ipva.text 

                if texto_ipva != "2025":

                    tabela_emissao_ipva = WebDriverWait(navegador, 60).until(
                        EC.visibility_of_element_located((By.CSS_SELECTOR, "div.forms-wrapper-body > div.forms-wrapper-body.forms-wrapper-body_table.pad-botton.ng-star-inserted > table > thead > tr"))
                    )   

                    botao_emitir_ipva = navegador.find_element(By.CSS_SELECTOR, "div:nth-child(3) > app-ipva > div > div.forms-wrapper-body > div.forms-wrapper-body.forms-wrapper-body_table.pad-botton.ng-star-inserted > table > tbody > tr:nth-child(7) > td.mat-cell.cdk-cell.cdk-column-A--o.mat-column-A--o.ng-star-inserted")  
                    botao_emitir_ipva.click()

                    time.sleep(8)

                    botao_ipva_execicio = WebDriverWait(navegador, 60).until(
                        EC.element_to_be_clickable((By.XPATH, "/html/body/app-root/div/div[1]/app-layout-servicos/app-debitos/app-card-detran/mat-card/mat-card-content/div[2]/app-ipva/div/div[2]/div[1]/p/mat-form-field/div/div[1]/div[3]/mat-select"))
                    ).click()

                    continue

                tabela_emissao_ipva = WebDriverWait(navegador, 60).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, "div.forms-wrapper-body > div.forms-wrapper-body.forms-wrapper-body_table.pad-botton.ng-star-inserted > table > thead > tr"))
                )

                botao_emitir_ipva = navegador.find_element(By.CSS_SELECTOR, "div:nth-child(3) > app-ipva > div > div.forms-wrapper-body > div.forms-wrapper-body.forms-wrapper-body_table.pad-botton.ng-star-inserted > table > tbody > tr:nth-child(1) > td.mat-cell.cdk-cell.cdk-column-A--o.mat-column-A--o.ng-star-inserted > button")
                botao_emitir_ipva.click()

                time.sleep(8)

                botao_ipva_execicio = WebDriverWait(navegador, 60).until(
                    EC.element_to_be_clickable((By.XPATH, "/html/body/app-root/div/div[1]/app-layout-servicos/app-debitos/app-card-detran/mat-card/mat-card-content/div[2]/app-ipva/div/div[2]/div[1]/p/mat-form-field/div/div[1]/div[3]/mat-select"))
                ).click()
            
            guia_dados[f"D{linhaPlan}"] = "OK!"

            botao_voltar_pesquisa = WebDriverWait(navegador, 60).until(
                EC.element_to_be_clickable((By.XPATH, "/html/body/app-root/div/div[1]/app-layout-servicos/app-debitos/app-card-detran/mat-card/mat-card-footer/mat-card-actions/button"))
            ).click()

        index +=1
        planilha.save(caminho_planilha)    

except Exception as e :

    print(f"Erro {e}")
    print("Renicie a automação")
    breakpoint()

navegador.quit()

arquivos_pdfs = [f for f in os.listdir(pasta_saida) if f.endswith('.pdf')]

for arquivo in arquivos_pdfs:

    caminho_pdf = os.path.join(pasta_saida, arquivo)

    with pdfplumber.open(caminho_pdf) as pdf:
        for num_page, pagina in enumerate(pdf.pages):

            texto_boleto = pagina.extract_text()
            print(texto_boleto)

            if "IPVA" in texto_boleto:

                corte_ipva_1 = re.split("17.Valor Total -", texto_boleto)
                valor_ipva = corte_ipva_1[2].replace(" ", "").replace("R$", "")

                corte_placa_1 = re.split("PLACA: ", texto_boleto)
                corte_placa_2 = re.split(" TIPO:", corte_placa_1[1])
                placa = corte_placa_2[0]

                corte_ano_ipva_1 = re.split(" IPVA ", texto_boleto)
                corte_ano_ipva_2 = re.split(" - DETRAN DIGITAL", corte_ano_ipva_1[1])
                ano_ipva_renomear = corte_ano_ipva_2[0]

                caminho_pdf_saida = os.path.join(pasta_saida, f"IPVA {placa}_{ano_ipva_renomear}.pdf")

                if "bordero" in arquivo:
                    if not os.path.exists(caminho_pdf_saida):
                        # Abre o PDF com PyPDF2
                        leitor_pdf = PdfReader(caminho_pdf)
                        # Criar um novo PDF para a página atual
                        escritor_pdf = PdfWriter()
                        escritor_pdf.add_page(leitor_pdf.pages[num_page])  # Adiciona a página correta


                        # Salva o PDF no caminho de saída
                        with open(caminho_pdf_saida, "wb") as arquivo_pdf_saida:
                            escritor_pdf.write(arquivo_pdf_saida)
                            print(f"Página salva no caminho: {caminho_pdf_saida}")
                            pdf.close()
                            os.remove(caminho_pdf)
                    else:
                        print(f"Arquivo já existe: {caminho_pdf_saida}")

                # Procurar a placa na planilha e preencher dados
                for linha in range(2, guia_dados.max_row + 1):
                    celula_placa = guia_dados[f'A{linha}'].value  # Coluna 'A' contém as placas
                    if celula_placa == placa: 
                        guia_dados[f'E{linha}'] = valor_ipva  # Coluna 'E': VALOR IPVA
                        planilha.save(caminho_planilha)
                        break
                    else:
                        print(f"Placa {placa} não encontrada na planilha.")
                
                continue
        
            corte_placa_1 = re.split("Placa: ", texto_boleto)
            corte_placa_2 = re.split(" Marca/Mod:", corte_placa_1[1])
            placa = corte_placa_2[0].replace("-", "")

            corte_valor_licenciamento_1 = re.split("ETC ", texto_boleto)
            corte_valor_licenciamento_2 = re.split("\nAutenticação Mecânica", corte_valor_licenciamento_1[1])
            valor_licenciamento = corte_valor_licenciamento_2[0].replace("R$", "").replace(" ", "")

            corte_propietario_1 = re.split("Proprietário: ", texto_boleto)
            corte_propietario_2 = re.split(" CPF: ", corte_propietario_1[1])
            propietario = corte_propietario_2[0]

            # Procurar a placa na planilha e preencher dados
            for linha in range(2, guia_dados.max_row + 1):
                celula_placa = guia_dados[f'A{linha}'].value  # Coluna 'A' contém as placas
                if celula_placa == placa:  # Verificar se a placa da planilha corresponde à do boleto
                    guia_dados[f'I{linha}'] = propietario  # Coluna 'B': PROPRIETARIO
                    guia_dados[f'G{linha}'] = valor_licenciamento  # Coluna 'D': VALOR LIC
                    guia_dados[f'J{linha}'] = "BOLETO PROCESSADO"  # Coluna 'F': STATUS BOLETO
                    print(f"Dados preenchidos para a placa {placa} na linha {linha}.")
                     # Salvar a planilha
                    planilha.save(caminho_planilha)
                    break
            else:
                print(f"Placa {placa} não encontrada na planilha.")

for linha in range (2, guia_dados.max_row + 1):
    
    celula_ipva = guia_dados[f'F{linha}'].value
    celula_licenciamento = guia_dados[f'G{linha}'].value
    
    # Verifica se algum dos valores é None ou está vazio
    if not celula_ipva or not celula_licenciamento:
        continue  # Pula para a próxima iteração se qualquer valor estiver vazio
    
    valor_ipva_calc = float(celula_ipva.replace('.', '').replace(',', '.'))
    valor_licenciamento_cal = float(celula_licenciamento.replace('.', '').replace(',', '.'))
    
    valor_total = valor_ipva_calc + valor_licenciamento_cal
    
    guia_dados[f'H{linha}'] = valor_total  # Coluna 'H': VALOR 
    planilha.save(caminho_planilha)
    
