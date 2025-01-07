#ROBO DE PESQUISA DE MULTAS
from openpyxl import load_workbook
import time
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import requests
from selenium.webdriver.common.keys import Keys  # Para acessar as teclas especiais
from selenium.webdriver.common.action_chains import ActionChains
import pyautogui
import undetected_chromedriver as uc
import requests
import os
import tkinter as tk
from tkinter import filedialog, messagebox
import threading
from dotenv import load_dotenv


# Variáveis globais para armazenar os caminhos selecionados
caminho_notificacao = ""
caminho_multas = ""
caminho_planilha = ""

"""
    Cria uma pasta de saída baseada no nome da planilha e em um caminho de base fornecido.
"""
def criar_pasta_saida(caminho_planilha, base_saida):
    """
    Cria uma pasta de saída baseada no nome da planilha e em um caminho de base fornecido.
    """
    # Pega o nome do arquivo da planilha sem a extensão
    nome_arquivo = os.path.splitext(os.path.basename(caminho_planilha))[0]
    
    # Caminho da pasta de saída com base na pasta base fornecida
    pasta_saida = os.path.join(base_saida, nome_arquivo)
    
    # Verifica se a pasta de saída já existe, se não existir, cria a pasta
    if not os.path.exists(pasta_saida):
        os.makedirs(pasta_saida)
        print(f"Pasta criada: {pasta_saida}")
    else:
        print(f"A pasta já existe: {pasta_saida}")
    
    # Cria as subpastas "multas" e "notificação" dentro da pasta de saída
    pasta_multas = os.path.join(pasta_saida, "multas")
    pasta_notificacao = os.path.join(pasta_saida, "notificacao")
    
    for pasta in [pasta_multas, pasta_notificacao]:
        if not os.path.exists(pasta):
            os.makedirs(pasta)
            print(f"Subpasta criada: {pasta}")
        else:
            print(f"A subpasta já existe: {pasta}")

    return pasta_saida, pasta_multas, pasta_notificacao

#Botão para selecionar a planilha excel para pesquisa
def selecionar_planilha():
    global caminho_planilha
    caminho_planilha = filedialog.askopenfilename(title="Selecione a Planilha")

def iniciar_automacao():
    try:

         # Simulação de um processo longo
        if not caminho_planilha:
            messagebox.showerror("Erro", "Por favor, selecione todos os arquivos antes de iniciar a automação.")
            return

        # Simulação de um processo longo
        messagebox.showinfo("Início", "A automação está sendo iniciada...")

        # Função que inicia a automação em uma thread separada
        pathSaida = r'M:\MULTAS\ROBO\Pesquisa BASE MG\Arquivos Processados'

        # Cria a pasta de saída com base na planilha selecionada
        pasta_saida, pasta_multas, pasta_notificacao = criar_pasta_saida(caminho_planilha, pathSaida)

        #Abri a planilha do Excel
        planilha = load_workbook(caminho_planilha)
        # Acessa as guias pelo nome
        guia_relacao_veiculos = planilha['RELACAO DE VEICULOS']
        guia_resultado_autuacao = planilha['RESULTADO AUTUACAO']
        guia_resultado_multa = planilha['RESULTADO MULTA']


        # Inicializa a variável linhaDaVez antes do loop
        cont = 1

        # Configurar o Chrome com um User-Agent falso usando undetected-chromedriver
        chrome_options = Options()
        chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36")

        # Inicializa o navegador com as opções e com o stealth mode ativado
        navegador = uc.Chrome(options=chrome_options)
        navegador.maximize_window()


        actions = ActionChains(navegador)


        # Sua chave API do 2Captcha
        API_KEY = "api_key"

        # data-sitekey do reCAPTCHA (obtido ao inspecionar a página)
        GOOGLE_KEY = '6LfVpnIUAAAAAHkISk6Z6juZcsUx6hbyJGwfnfPL'

        # URL da página com o reCAPTCHA
        PAGE_URL = 'https://www.transito.mg.gov.br/veiculos/situacao-do-veiculo/consultar-situacao-do-veiculo/'

        # 1. Enviar requisição para o 2Captcha para resolver o reCAPTCHA
        def enviar_requisicao_captcha(api_key, google_key, page_url):
            url = f"http://2captcha.com/in.php?key={api_key}&method=userrecaptcha&googlekey={google_key}&pageurl={page_url}"
            response = requests.get(url)
            if response.status_code == 200 and 'OK|' in response.text:
                captcha_id = response.text.split('|')[1]
                return captcha_id
            else:
                raise Exception(f"Erro ao enviar captcha: {response.text}")

        # 2. Obter o token de resposta do captcha
        def obter_resposta_captcha(api_key, captcha_id):
            url = f"http://2captcha.com/res.php?key={api_key}&action=get&id={captcha_id}"
            while True:
                response = requests.get(url)
                if 'CAPCHA_NOT_READY' in response.text:
                    print("Captcha ainda não resolvido, tentando novamente em 5 segundos...")
                    time.sleep(5)
                elif 'OK|' in response.text:
                    return response.text.split('|')[1]
                else:
                    raise Exception(f"Erro ao obter resposta do captcha: {response.text}")

        # 3. Inserir o token na página (use Selenium ou outro método)
        def inserir_token(token):
            
            # Espera o carregamento da página e insere o token
            navegador.execute_script(f"document.getElementById('g-recaptcha-response').innerHTML = '{token}'")

        navegador.get("https://transito.mg.gov.br/veiculos/situacao-do-veiculo/consultar-situacao-do-veiculo")

        #Função para contornar erro

        def erroSite():
            try:               
                #Valida a tela de erro
                telaErro1 = WebDriverWait(navegador,5).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, "body > main > div > div.card > div > div > div > div > h2"))
                )
                navegador.refresh()
                
                pyautogui.hotkey('f5')   
                                
                time.sleep(1)
                
                pyautogui.hotkey('enter')
                
                time.sleep(10)
                                            
            except TimeoutException:
                print("Prosseguir cadastro")

        try:
            
            #ESPERA A TELA PARA REALZIAR O LOGIN NO GOV
            telaParaLoginGov = WebDriverWait(navegador, 15).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, "body > main > div > div.row"))
            )
        except TimeoutException:
            
            #SE DER ERRO NAVEGA PARA A PAGINA NOVAMENTE
            navegador.get("https://www.transito.mg.gov.br/veiculos/situacao-do-veiculo/consultar-situacao-do-veiculo/")
            
        #CLICA NO BOTÃO PARA REALIZAR O LOGIN COM O GOV
        botaoLoginGov = navegador.find_element(By.CSS_SELECTOR, "#content > div > div > div.row.text-center.h5.mt-2 > div > a")
        navegador.execute_script("arguments[0].click();", botaoLoginGov)

        #Espera a primeira tela do login do GOV
        tela1GOV = WebDriverWait(navegador, 100).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "body > div.container"))
        )

        # Cria uma instância de ActionChains
        actions = ActionChains(navegador)

        cpfLogin = "CPF"
        senhaLogin = "SENHA"

        #Inseri o CPF no campo do CPF
        campoCPF = navegador.find_element(By.CSS_SELECTOR, "#accountId")
        for caractere in cpfLogin:
            campoCPF.send_keys(caractere)
            time.sleep(0.2)
            

        #Clica no botão para prosseguir para outra pagina para colocar a senha
        botaoContinuarLogin = navegador.find_element(By.CSS_SELECTOR, "#enter-account-id")
        navegador.execute_script("arguments[0].click();", botaoContinuarLogin)

        #Espera a segunda tela de login do GOV
        tela2GOV = WebDriverWait(navegador, 1000).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "#login-password-info > img"))
        )


        #Inseri a senha no site
        campoSENHA = navegador.find_element(By.CSS_SELECTOR, "#password")
        for caractere in senhaLogin:
            campoSENHA.send_keys(caractere)
            time.sleep(0.2)


        #Entra no site
        botaoENTRAR = navegador.find_element(By.CSS_SELECTOR, "#submit-button")
        navegador.execute_script("arguments[0].click();", botaoENTRAR)

        try:
            
            #Espera a pagina pricipal do DETRAN MG
            paginaPrincipal = WebDriverWait(navegador, 30).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, "body div.col-sm.col.px-0"))
            )
            
            #Carrega os cabeçalhos da planilha
            guia_relacao_veiculos['A1'] = "Status"
            guia_relacao_veiculos['B1'] = "Placa"
            guia_relacao_veiculos['C1'] = "Chassi"
            guia_relacao_veiculos['D1'] = "RENAVAM"
            guia_relacao_veiculos['E1'] = "Situação"
            guia_relacao_veiculos['F1'] = "Qtde Autuação"
            guia_relacao_veiculos['G1'] = "Qtde Multas"
            guia_relacao_veiculos['H1'] = "OBS"
            guia_relacao_veiculos['I1'] = "conf. Autuação"
            guia_relacao_veiculos['J1'] = "conf. Multas"
            
            #Passa os valores para uma lista
            status = []
            placa = []
            chassi = []
            renavam = []
            
            time.sleep(2)
            

            linhaResult = 1
                
            for index, row in enumerate(guia_relacao_veiculos.iter_rows(min_row=2, max_row=guia_relacao_veiculos.max_row), start=0):
                
                numerosMaxLinhas = guia_relacao_veiculos.max_row
                if (index == numerosMaxLinhas -1 ):
                    print("Pesquisa de placas finalizadas")
                    break
                
                linhaResult +=1
                status.append(row[0].value) #Obtem o status
                placa.append(row[1].value) #Obtem o valor da placa
                chassi.append(row[2].value) #Obtem o chassi
                renavam.append(row[3].value) #Obtem renavam
                
                if status[index] is None:
                    
                    #Inseri a placa no campo placa
                    campoPlaca = navegador.find_element(By.CSS_SELECTOR, "#placa")
                    campoPlaca.clear()
                    campoPlaca.send_keys(placa[index])
                    
                    #Inseri o chassi no campo chassi
                    campoChassi = navegador.find_element(By.CSS_SELECTOR, "#chassi")
                    campoChassi.clear()
                    campoChassi.send_keys(chassi[index])
                    
                    # Função principal
                    def resolver_recaptcha(tentativas = 0, max_Tentativas = 3):
                        
                        if tentativas >= max_Tentativas:
                            print("Número maximo de tentativas atingido. Pesquisando próximo veículo")
                            # Marca como "Erro ao resolver CAPTCHA" na planilha, e avança para o próximo
                            guia_relacao_veiculos[f'A{linhaResult}'] = "Erro ao pesquisar"
                            planilha.save(caminho_planilha)
                            return  # Retorna para continuar o loop principal
                        
                        try:
                            print("Enviando requisição para resolver reCAPTCHA...")
                            captcha_id = enviar_requisicao_captcha(API_KEY, GOOGLE_KEY, PAGE_URL)
                            
                            print("Aguardando solução do captcha...")
                            token = obter_resposta_captcha(API_KEY, captcha_id)
                            
                            print("Inserindo token e enviando formulário...")
                            inserir_token(token)
                            
                            print("reCAPTCHA resolvido com sucesso!")
                            
                            botaoPEsquisar = navegador.find_element(By.CSS_SELECTOR, "#content > form > button")
                            navegador.execute_script("arguments[0].click();", botaoPEsquisar)
                        
                            #=============================================================================================
                            
                            #Passa pelo erro do site
                            erroSite()
                            
                            #Verifica se o veículo não esta cadastrado na base Minas
                            try:
                                
                                campoVeiculoNaoCadastrado = navegador.find_element(By.CSS_SELECTOR, "div.alert.alert-danger.px-3.py-2.font-weight-bold.h5")
                                valorVeiculoNaoCadastrado = campoVeiculoNaoCadastrado.text
                                
                                if (valorVeiculoNaoCadastrado == 'Veículo não Cadastrado na Base de Minas Gerais'):
                                    guia_relacao_veiculos[f'H{linhaResult}'] = "Veículo não Cadastrado na Base de Minas Gerais"  
                                    planilha.save(caminho_planilha)
                                    
                                    #Retorna para buscar o proximo veiculo
                                    campoConsultaOutroVeiculo = WebDriverWait(navegador, 15).until(
                                        EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.text-secondary[href*="/veiculos/situacao-do-veiculo/consultar-situacao-do-veiculo/"]'))
                                    ).click()
                                    
                                    #Coloca o status na planilha
                                    guia_relacao_veiculos[f'A{linhaResult}'] = "OK!"  # Inseri o status na planilha
                                    planilha.save(caminho_planilha)
                                            
                                    return  # Retorna da função para continuar o loop principal

                            except NoSuchElementException:
                                
                                print("Veículo cadastrado na base Minas")
                                
                            time.sleep(3)
                            
                            #                               EMISSÃO DE NOTIFICAÇÃO
                            
                            try:
                                
                                #Botão de autuações
                                campoAutuacao = WebDriverWait(navegador, 15).until(
                                    EC.visibility_of_element_located((By.CSS_SELECTOR, "#content > h2:nth-child(6)"))
                                )
                                campoAutuacao.click()
                                
                                #Esperar o modal da autuação abrir 
                                modalAutuacao = WebDriverWait(navegador, 15).until(
                                    EC.visibility_of_element_located((By.CSS_SELECTOR, "#divDefesas"))
                                )
                                
                                time.sleep(1)
                                
                                # Localizando a tabela pela classe
                                tabelaAtuacao = navegador.find_element(By.CSS_SELECTOR, ".table.table-sm.table-striped.table-bordered")
                                
                                # Contando o número de linhas (tr)
                                linhasAutuacao = tabelaAtuacao.find_elements(By.TAG_NAME, "tr")
                                numero_linhas = len(linhasAutuacao)
                                
                                quantidadeAutuacoes = 0  # Inicializa a variável para acumular as autuações
                                
                                time.sleep(2)
                                
                                
                                #EMITINDO AS NOTIFICAÕES
                                for cont in range(2, numero_linhas + 1):
                                    
                                    #Seleciona a quantidade de multas
                                    seletorQuantidadeAutuacoes = f"#divDefesas > div > table > tbody > tr:nth-child({cont}) > td.text-center"
                                    campoQuantidadeAutuacoes = navegador.find_element(By.CSS_SELECTOR, seletorQuantidadeAutuacoes)
                                    
                                    # Captura o texto da célula
                                    valor = campoQuantidadeAutuacoes.text
                                    
                                    valor = int(valor)
                                    
                                    quantidadeAutuacoes += valor
                                    
                                    time.sleep(2)

                                    if (valor == 1):
                                        
                                        #Clica na multa atual para realizar a emissão
                                        seletorEmissaoAutuacao = f"#divDefesas > div > table > tbody > tr:nth-child({cont}) > td > a"
                                        campoEmissaoAutuacao = navegador.find_element(By.CSS_SELECTOR, seletorEmissaoAutuacao)
                                        navegador.execute_script("arguments[0].click();", campoEmissaoAutuacao)
                                    
                                        #Pula o erro do site                                erroSite()
                                        
                                        esperaAutuacao  = WebDriverWait(navegador, 10).until(
                                            EC.visibility_of_element_located((By.CSS_SELECTOR, "#content > dl > dd:nth-child(2)"))
                                        )
                                        
                                        #Obtem o orgão autuador
                                        campoOrgaoAutuador = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(2)")
                                        valorOrgaoAutuador = campoOrgaoAutuador.text
                                        
                                        #Obtem a sistuação
                                        campoSituacao = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(4)")
                                        valorSituacao = campoSituacao.text
                                        
                                        #Obtem o codigo da multa
                                        campoCodigoMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(8)")
                                        valorCodigoMulta = campoCodigoMulta.text
                                        
                                        #Obtem a data da multa
                                        campoDataMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(10)")
                                        valorDataMulta = campoDataMulta.text
                                        
                                        #Obtem a hora da multa
                                        campoHoraMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(12)")
                                        valorHoraMulta = campoHoraMulta.text
                                        
                                        #Obtem a descrição da multa
                                        campoDescricaoMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(14)") 
                                        valorDescricaoMulta = campoDescricaoMulta.text
                                        
                                        #obtem o local da multa 
                                        campoLocalMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(16)")
                                        valorLocalMulta = campoLocalMulta.text
                                        
                                        #Obtem o municipio 
                                        campoMunicipioMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(18)")
                                        valorMunicipioMulta = campoMunicipioMulta.text
                                        
                                        #Obtem a incluida em
                                        campoIncluidaMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(20)")
                                        valorIncluidaMulta = campoIncluidaMulta.text
                                                                
                                        #Obtem a data limite defesa
                                        campoDataLimiteDefesa = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(22)")
                                        valorDataLimiteDefesa = campoDataLimiteDefesa.text
                                        
                                        #Obtem o numero do AIT
                                        campoNumeroAIT = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(24)")
                                        valorNumeroAIT = campoNumeroAIT.text
                                        
                                        #Obtem o numero do processamento
                                        campoNumeroProcessamento = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(26)")
                                        valorNumeroProcessamento = campoNumeroProcessamento.text
                                        
                                        # Converte os valores da coluna B em um conjunto para facilitar a busca
                                        colunaAUTO = set(cell.value for cell in guia_resultado_autuacao['B'] if cell.value is not None)

                                        # Valor que você quer verificar
                                        valorAITProcurar = valorNumeroAIT

                                        # Define `existe` como True se o valor estiver no conjunto, caso contrário, False
                                        existe = valorAITProcurar in colunaAUTO
                    
                                        if not existe:

                                            ultimaLinha1 = guia_resultado_autuacao.max_row + 1
                                                
                                            # INSERINDO OS VALORES OBTIDOS DAS AUTUAÇÕES NA PLANILHA
                                            guia_resultado_autuacao[f'A{ultimaLinha1}'] = placa[index]  # Inserindo o valor da placa
                                            guia_resultado_autuacao[f'B{ultimaLinha1}'] = valorNumeroAIT  # Inserindo o número do AIT
                                            guia_resultado_autuacao[f'C{ultimaLinha1}'] = valorDataMulta  # Inserindo o valor da data da multa
                                            guia_resultado_autuacao[f'D{ultimaLinha1}'] = valorHoraMulta  # Inserindo a hora da multa
                                            guia_resultado_autuacao[f'E{ultimaLinha1}'] = valorCodigoMulta  # Inserindo o valor do código da multa
                                            guia_resultado_autuacao[f'F{ultimaLinha1}'] = valorNumeroProcessamento  # Inserindo o número de processamento
                                            guia_resultado_autuacao[f'G{ultimaLinha1}'] = valorSituacao  # Inserindo o valor da situação
                                            guia_resultado_autuacao[f'H{ultimaLinha1}'] = valorDescricaoMulta  # Inserindo o valor da descrição da multa
                                            guia_resultado_autuacao[f'I{ultimaLinha1}'] = valorLocalMulta  # Inserindo o valor do local da multa
                                            guia_resultado_autuacao[f'J{ultimaLinha1}'] = valorMunicipioMulta  # Inserindo o valor do município na multa
                                            guia_resultado_autuacao[f'K{ultimaLinha1}'] = valorIncluidaMulta  # Inserindo o valor de quando a multa é incluída
                                            guia_resultado_autuacao[f'L{ultimaLinha1}'] = valorDataLimiteDefesa  # Inserindo o valor da data limite da multa
                                            guia_resultado_autuacao[f'M{ultimaLinha1}'] = valorOrgaoAutuador  # Inserindo o valor do órgão autuador
                                            
                                            #Salva os resultados na planilha
                                            planilha.save(caminho_planilha)
                                            
                                            time.sleep(3)

                                            #EMITI A NOTIFICAÇÃO
                                            pyautogui.hotkey('ctrl', 'p')
                                            time.sleep(5)

                                            # Caminha para selecionar o download como PDF
                                            for i in range(5):
                                                pyautogui.hotkey('tab')
                                            time.sleep(1.5)
                                            pyautogui.write("salvar")
                                            time.sleep(1.5)
                                            for i in range(6):
                                                pyautogui.hotkey('tab')
                                            time.sleep(1.5)
                                            pyautogui.hotkey('enter')

                                            time.sleep(3)
                                            # Define o caminho de salvamento
                                            caminhoAutuacao = os.path.join(pasta_notificacao, valorNumeroAIT)

                                            # Torna o caminho compatível com digitação em `pyautogui`
                                            caminhoAutuacao = os.path.normpath(caminhoAutuacao)  # Converte barras para o formato correto
                                            pyautogui.write(caminhoAutuacao)
                                            time.sleep(1)
                                            pyautogui.hotkey('enter')     

                                        #Clica para consulta a proxima autuação
                                        campoVoltarDadosVeiculos = navegador.find_element(By.CSS_SELECTOR, "#content > div:nth-child(8) > a")
                                        navegador.execute_script("arguments[0].click();", campoVoltarDadosVeiculos)
                                        
                                        time.sleep(1)
                                        
                                        #Valida a tela de erro
                                        erroSite()
                                        
                                        #Botão de autuações
                                        campoAutuacao = WebDriverWait(navegador, 15).until(
                                                EC.visibility_of_element_located((By.CSS_SELECTOR, "#content > h2:nth-child(6)"))
                                        )
                                        campoAutuacao.click()
                                        
                                        time.sleep(2)
                                        
                                    else:
                                        
                                        time.sleep(2)
                                        
                                        #Clica na multa atual para realizar a emissão
                                        seletorEmissaoAutuacao = f"#divDefesas > div > table > tbody > tr:nth-child({cont}) > td > a"
                                        campoEmissaoAutuacao = navegador.find_element(By.CSS_SELECTOR, seletorEmissaoAutuacao)
                                        navegador.execute_script("arguments[0].click();", campoEmissaoAutuacao)

                                        #Valida a tela de erro
                                        erroSite()
                                        
                                        esperaAutuacao  = WebDriverWait(navegador, 10).until(
                                            EC.visibility_of_element_located((By.CSS_SELECTOR, "#content > dl > dd:nth-child(2)"))
                                        )
                                        
                                        for i in range(0, valor):
                                            
                                            #Valida a tela de erro
                                            erroSite()    
                                            
                                            #Obtem o orgão autuador
                                            campoOrgaoAutuador = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(2)")
                                            valorOrgaoAutuador = campoOrgaoAutuador.text
                                            
                                            #Obtem a sistuação
                                            campoSituacao = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(4)")
                                            valorSituacao = campoSituacao.text
                                            
                                            #Obtem o codigo da multa
                                            campoCodigoMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(8)")
                                            valorCodigoMulta = campoCodigoMulta.text
                                            
                                            #Obtem a data da multa
                                            campoDataMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(10)")
                                            valorDataMulta = campoDataMulta.text
                                            
                                            #Obtem a hora da multa
                                            campoHoraMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(12)")
                                            valorHoraMulta = campoHoraMulta.text
                                            
                                            #Obtem a descrição da multa
                                            campoDescricaoMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(14)") 
                                            valorDescricaoMulta = campoDescricaoMulta.text
                                            
                                            #obtem o local da multa 
                                            campoLocalMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(16)")
                                            valorLocalMulta = campoLocalMulta.text
                                            
                                            #Obtem o municipio 
                                            campoMunicipioMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(18)")
                                            valorMunicipioMulta = campoMunicipioMulta.text
                                            
                                            #Obtem a incluida em
                                            campoIncluidaMulta = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(20)")
                                            valorIncluidaMulta = campoIncluidaMulta.text
                                                                    
                                            #Obtem a data limite defesa
                                            campoDataLimiteDefesa = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(22)")
                                            valorDataLimiteDefesa = campoDataLimiteDefesa.text
                                            
                                            #Obtem o numero do AIT
                                            campoNumeroAIT = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(24)")
                                            valorNumeroAIT = campoNumeroAIT.text
                                            
                                            #Obtem o numero do processamento
                                            campoNumeroProcessamento = navegador.find_element(By.CSS_SELECTOR, "#content > dl > dd:nth-child(26)")
                                            valorNumeroProcessamento = campoNumeroProcessamento.text
                                            
                                            # Converte os valores da coluna B em um conjunto para facilitar a busca
                                            colunaAUTO = set(cell.value for cell in guia_resultado_autuacao['B'] if cell.value is not None)

                                            # Valor que você quer verificar
                                            valorAITProcurar = valorNumeroAIT

                                            # Define `existe` como True se o valor estiver no conjunto, caso contrário, False
                                            existe = valorAITProcurar in colunaAUTO
                    
                                            if not existe:
                                                
                                                ultimaLinha2 = guia_resultado_autuacao.max_row + 1
                                                    
                                                # INSERINDO OS VALORES OBTIDOS DAS AUTUAÇÕES NA PLANILHA
                                                guia_resultado_autuacao[f'A{ultimaLinha2}'] = placa[index]  # Inserindo o valor da placa
                                                guia_resultado_autuacao[f'B{ultimaLinha2}'] = valorNumeroAIT  # Inserindo o número do AIT
                                                guia_resultado_autuacao[f'C{ultimaLinha2}'] = valorDataMulta  # Inserindo o valor da data da multa
                                                guia_resultado_autuacao[f'D{ultimaLinha2}'] = valorHoraMulta  # Inserindo a hora da multa
                                                guia_resultado_autuacao[f'E{ultimaLinha2}'] = valorCodigoMulta  # Inserindo o valor do código da multa
                                                guia_resultado_autuacao[f'F{ultimaLinha2}'] = valorNumeroProcessamento  # Inserindo o número de processamento
                                                guia_resultado_autuacao[f'G{ultimaLinha2}'] = valorSituacao  # Inserindo o valor da situação
                                                guia_resultado_autuacao[f'H{ultimaLinha2}'] = valorDescricaoMulta  # Inserindo o valor da descrição da multa
                                                guia_resultado_autuacao[f'I{ultimaLinha2}'] = valorLocalMulta  # Inserindo o valor do local da multa
                                                guia_resultado_autuacao[f'J{ultimaLinha2}'] = valorMunicipioMulta  # Inserindo o valor do município na multa
                                                guia_resultado_autuacao[f'K{ultimaLinha2}'] = valorIncluidaMulta  # Inserindo o valor de quando a multa é incluída
                                                guia_resultado_autuacao[f'L{ultimaLinha2}'] = valorDataLimiteDefesa  # Inserindo o valor da data limite da multa
                                                guia_resultado_autuacao[f'M{ultimaLinha2}'] = valorOrgaoAutuador  # Inserindo o valor do órgão autuador
                                                
                                                #Salva os resultados na planilha
                                                planilha.save(caminho_planilha)
                                                
                                                time.sleep(3)

                                                #EMITI A NOTIFICAÇÃO
                                                pyautogui.hotkey('ctrl', 'p')
                                                time.sleep(5)

                                                # Caminha para selecionar o download como PDF
                                                for i in range(5):
                                                    pyautogui.hotkey('tab')
                                                time.sleep(1.5)
                                                pyautogui.write("salvar")
                                                time.sleep(1.5)
                                                for i in range(6):
                                                    pyautogui.hotkey('tab')
                                                time.sleep(1.5)
                                                pyautogui.hotkey('enter')

                                                time.sleep(3)
                                                # Define o caminho de salvamento
                                                caminhoAutuacao = os.path.join(pasta_notificacao, valorNumeroAIT)

                                                # Torna o caminho compatível com digitação em `pyautogui`
                                                caminhoAutuacao = os.path.normpath(caminhoAutuacao)  # Converte barras para o formato correto
                                                pyautogui.write(caminhoAutuacao)
                                                time.sleep(1)
                                                pyautogui.hotkey('enter') 
                                            
                                            try:
                                                
                                                #Clica no botão proxima autuação
                                                botaoProxAutuacao = navegador.find_element(By.XPATH, '//*[@id="content"]/div[3]/div/a')
                                                navegador.execute_script("arguments[0].click();", botaoProxAutuacao)
                                            except:
                                                continue
                                            
                                        #Clica para consulta a proxima autuação
                                        campoVoltarDadosVeiculos = navegador.find_element(By.CSS_SELECTOR, "#content > div:nth-child(8) > a")
                                        navegador.execute_script("arguments[0].click();", campoVoltarDadosVeiculos)
                                        
                                        time.sleep(2)
                                        
                                        #Valida a tela de erro
                                        erroSite()
                                        
                                        
                                        #Botão de autuações
                                        campoAutuacao = WebDriverWait(navegador, 15).until(
                                                EC.visibility_of_element_located((By.CSS_SELECTOR, "#content > h2:nth-child(6)"))
                                        )
                                        campoAutuacao.click()
                                        
                                        time.sleep(2)

                                time.sleep(2)
                                guia_relacao_veiculos[f'F{linhaResult}'] = quantidadeAutuacoes  # Inserindo a quantidade de autuações
                                planilha.save(caminho_planilha)
                            
                            except TimeoutException:
                                guia_relacao_veiculos[f'F{linhaResult}'] = 0  # Inserindo a quantidade de autuações
                                planilha.save(caminho_planilha)
                                print("Veículo sem notificação")
                                
                            #limpa a variavel numero_linhas
                            numero_linhas = 0
                            valor = 0
                            
                            
                            #                               EMISSÃO DE MULTAS
                            
                            try:
                                
                                #Espera a janela de multas
                                campoMultas = WebDriverWait(navegador, 15).until(
                                    EC.visibility_of_element_located((By.CSS_SELECTOR, "#content > h2:nth-child(8)"))
                                ).click()
                                
                                #Esperar o modal da autuação abrir 
                                modalMultas = WebDriverWait(navegador, 15).until(
                                    EC.visibility_of_element_located((By.CSS_SELECTOR, "#divMultas"))
                                )
                                
                                time.sleep(1)
                                
                                tabelaMultas = WebDriverWait(navegador, 10).until(
                                    EC.presence_of_element_located((By.CSS_SELECTOR, "#divMultas > div > table"))
                                )
                                
                                
                                # Contando o número de linhas (tr)
                                linhasMultas = tabelaMultas.find_elements(By.TAG_NAME, "tr")
                                numero_linhas = len(linhasMultas)
                                

                                # Inicializa a variável para acumular as autuações
                                quantidadeMultas = 0  
                                
                                time.sleep(1.5)
                                
                                for cont in range(2, numero_linhas + 1):  
                                    #Seleciona a quantidade de multas
                                    seletorQuantidadeMultas = f"#divMultas > div > table > tbody > tr:nth-child({cont}) > td.text-center"
                                    campoQuantidadeMultas = navegador.find_element(By.CSS_SELECTOR, seletorQuantidadeMultas)
                                    
                                    # Captura o texto da célula
                                    valor = campoQuantidadeMultas.text
                                    
                                    valor = int(valor)
                                    
                                    quantidadeMultas += valor
                                    
                                    time.sleep(2)
                                    
                                    if (valor == 1):
                                        
                                        #Clica na multa atual para realizar a emissão
                                        seletorEmissaoMulta = f"#divMultas > div > table > tbody > tr:nth-child({cont}) > td > a"
                                        campoEmissaoMulta = navegador.find_element(By.CSS_SELECTOR, seletorEmissaoMulta)
                                        navegador.execute_script("arguments[0].click();", campoEmissaoMulta)
                                    
                                        #Valida a tela de erro
                                        erroSite()
                                        
                                        esperaMulta  = WebDriverWait(navegador, 10).until(
                                            EC.visibility_of_element_located((By.CSS_SELECTOR, "#content dl:nth-child(1) > dd"))
                                        )
                                        
                                        #Obtem o orgão autuador
                                        campoOrgaoAutuador = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(1) > dd")
                                        valorOrgaoAutuador = campoOrgaoAutuador.text
                                            
                                        #Obtem a sistuação
                                        campoSituacao = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(2) > dd")
                                        valorSituacao = campoSituacao.text
                                            
                                        #Obtem o codigo da multa
                                        campoCodigoMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(4) > dd")
                                        valorCodigoMulta = campoCodigoMulta.text
                                            
                                        #Obtem a data da multa
                                        campoDataMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(5) > dd")
                                        valorDataMulta = campoDataMulta.text
                                            
                                        #Obtem a hora da multa
                                        campoHoraMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(6) > dd")
                                        valorHoraMulta = campoHoraMulta.text
                                            
                                        #Obtem a descrição da multa
                                        campoDescricaoMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(7) > dd") 
                                        valorDescricaoMulta = campoDescricaoMulta.text
                                            
                                        #obtem o local da multa 
                                        campoLocalMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(8) > dd")
                                        valorLocalMulta = campoLocalMulta.text
                                            
                                        #Obtem o municipio 
                                        campoMunicipioMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(9) > dd")
                                        valorMunicipioMulta = campoMunicipioMulta.text
                                            
                                        #Obtem a incluida em
                                        campoIncluidaMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(10) > dd")
                                        valorIncluidaMulta = campoIncluidaMulta.text
                                                                    
                                        #Obtem a data limite defesa
                                        campoDataLimiteDefesa = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(11) > dd")
                                        valorDataLimiteDefesa = campoDataLimiteDefesa.text
                                            
                                        #Obtem o numero do AIT
                                        campoNumeroAIT = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(12) > dd")
                                        valorNumeroAIT = campoNumeroAIT.text
                                            
                                        #Obtem o numero do processamento
                                        campoNumeroProcessamento = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(13) > dd")
                                        valorNumeroProcessamento = campoNumeroProcessamento.text
                                        
                                        #Obtem o valor da multa
                                        campoValorMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(14) > dd")
                                        valorMulta = campoValorMulta.text
                                            
                                        # Converte os valores da coluna B em um conjunto para facilitar a busca
                                        colunaAUTO = set(cell.value for cell in guia_resultado_multa['B'] if cell.value is not None)

                                        # Valor que você quer verificar
                                        valorAITProcurar = valorNumeroAIT

                                        # Define `existe` como True se o valor estiver no conjunto, caso contrário, False
                                        existe = valorAITProcurar in colunaAUTO
                    
                                        if not existe:
                                                
                                            ultimaLinha3 = guia_resultado_multa.max_row + 1
                                            
                                            # INSERINDO OS VALORES OBTIDOS DAS MULTAS NA PLANILHA
                                            guia_resultado_multa[f'A{ultimaLinha3}'] = placa[index]  # Inserindo o valor da placa
                                            guia_resultado_multa[f'B{ultimaLinha3}'] = renavam[index]  # Inserindo o valor do renavam
                                            guia_resultado_multa[f'C{ultimaLinha3}'] = valorNumeroAIT  # Inserindo o número do AIT
                                            guia_resultado_multa[f'D{ultimaLinha3}'] = valorDataMulta  # Inserindo o valor da data da multa
                                            guia_resultado_multa[f'E{ultimaLinha3}'] = valorHoraMulta  # Inserindo a hora da multa
                                            guia_resultado_multa[f'F{ultimaLinha3}'] = valorCodigoMulta  # Inserindo o valor do código da multa
                                            guia_resultado_multa[f'G{ultimaLinha3}'] = valorNumeroProcessamento  # Inserindo o número de processamento
                                            guia_resultado_multa[f'H{ultimaLinha3}'] = valorSituacao  # Inserindo o valor da situação
                                            guia_resultado_multa[f'I{ultimaLinha3}'] = valorDescricaoMulta  # Inserindo o valor da descrição da multa
                                            guia_resultado_multa[f'J{ultimaLinha3}'] = valorLocalMulta  # Inserindo o valor do local da multa
                                            guia_resultado_multa[f'K{ultimaLinha3}'] = valorMunicipioMulta  # Inserindo o valor do município na multa
                                            guia_resultado_multa[f'L{ultimaLinha3}'] = valorIncluidaMulta  # Inserindo o valor de quando a multa é incluída
                                            guia_resultado_multa[f'M{ultimaLinha3}'] = valorDataLimiteDefesa  # Inserindo o valor da data limite da multa
                                            guia_resultado_multa[f'N{ultimaLinha3}'] = valorOrgaoAutuador  # Inserindo o valor do órgão autuador
                                            guia_resultado_multa[f'O{ultimaLinha3}'] = valorMulta  # Inserindo o valor da multa-
                                            
                                            #Salva os resultados na planilha
                                            planilha.save(caminho_planilha)
                                            
                                            time.sleep(2)
                                    
                                        #Clica para consulta a proxima multa
                                        campoVoltarDadosVeiculos = navegador.find_element(By.CSS_SELECTOR, "#content > div:nth-child(8) > a")
                                        navegador.execute_script("arguments[0].click();", campoVoltarDadosVeiculos)
                                        
                                        time.sleep(1)
                                        
                                        #Valida a tela de erro
                                        erroSite()
                                                    
                                        #Botão de autuações
                                        campoMulta = WebDriverWait(navegador, 15).until(
                                                EC.visibility_of_element_located((By.CSS_SELECTOR, "#content > h2:nth-child(8)"))
                                        )
                                        campoMulta.click()
                                        
                                        time.sleep(2)
                                        
                                    else:
                                        
                                        time.sleep(2)
                                        #Clica na multa atual para realizar a emissão
                                        seletorEmissaoMulta = f"#divMultas > div > table > tbody > tr:nth-child({cont}) > td > a"
                                        campoEmissaoMulta = navegador.find_element(By.CSS_SELECTOR, seletorEmissaoMulta)
                                        navegador.execute_script("arguments[0].click();", campoEmissaoMulta)

                                        #Valida a tela de erro
                                        erroSite()    
                                            
                                        esperaMulta  = WebDriverWait(navegador, 10).until(
                                            EC.visibility_of_element_located((By.CSS_SELECTOR, "#content  dl:nth-child(1) > dt"))
                                        )
                                        
                                        for i in range(0, valor):
                                            
                                            #Valida a tela de erro
                                            erroSite()    
                                            
                                            #Obtem o orgão autuador
                                            campoOrgaoAutuador = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(1) > dd")
                                            valorOrgaoAutuador = campoOrgaoAutuador.text
                                            
                                            #Obtem a sistuação
                                            campoSituacao = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(2) > dd")
                                            valorSituacao = campoSituacao.text
                                            
                                            #Obtem o codigo da multa
                                            campoCodigoMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(4) > dd")
                                            valorCodigoMulta = campoCodigoMulta.text
                                            
                                            #Obtem a data da multa
                                            campoDataMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(5) > dd")
                                            valorDataMulta = campoDataMulta.text
                                            
                                            #Obtem a hora da multa
                                            campoHoraMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(6) > dd")
                                            valorHoraMulta = campoHoraMulta.text
                                            
                                            #Obtem a descrição da multa
                                            campoDescricaoMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(7) > dd") 
                                            valorDescricaoMulta = campoDescricaoMulta.text
                                            
                                            #obtem o local da multa 
                                            campoLocalMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(8) > dd")
                                            valorLocalMulta = campoLocalMulta.text
                                            
                                            #Obtem o municipio 
                                            campoMunicipioMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(9) > dd")
                                            valorMunicipioMulta = campoMunicipioMulta.text
                                            
                                            #Obtem a incluida em
                                            campoIncluidaMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(10) > dd")
                                            valorIncluidaMulta = campoIncluidaMulta.text
                                                                    
                                            #Obtem a data limite defesa
                                            campoDataLimiteDefesa = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(11) > dd")
                                            valorDataLimiteDefesa = campoDataLimiteDefesa.text
                                            
                                            #Obtem o numero do AIT
                                            campoNumeroAIT = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(12) > dd")
                                            valorNumeroAIT = campoNumeroAIT.text
                                            
                                            #Obtem o numero do processamento
                                            campoNumeroProcessamento = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(13) > dd")
                                            valorNumeroProcessamento = campoNumeroProcessamento.text
                                            
                                            #Obtem o valor da multa
                                            campoValorMulta = navegador.find_element(By.CSS_SELECTOR, "#content dl:nth-child(14) > dd")
                                            valorMulta = campoValorMulta.text
                                                
                                            # Converte os valores da coluna B em um conjunto para facilitar a busca
                                            colunaAUTO = set(cell.value for cell in guia_resultado_multa['B'] if cell.value is not None)

                                            # Valor que você quer verificar
                                            valorAITProcurar = valorNumeroAIT

                                            # Define `existe` como True se o valor estiver no conjunto, caso contrário, False
                                            existe = valorAITProcurar in colunaAUTO
                            
                                            if not existe:
                                                
                                                ultimaLinha4 = guia_resultado_multa.max_row + 1
                                                
                                                # INSERINDO OS VALORES OBTIDOS DAS MULTAS NA PLANILHA
                                                guia_resultado_multa[f'A{ultimaLinha4}'] = placa[index]  # Inserindo o valor da placa
                                                guia_resultado_multa[f'B{ultimaLinha4}'] = renavam[index]  # Inserindo o valor do renavam
                                                guia_resultado_multa[f'C{ultimaLinha4}'] = valorNumeroAIT  # Inserindo o número do AIT
                                                guia_resultado_multa[f'D{ultimaLinha4}'] = valorDataMulta  # Inserindo o valor da data da multa
                                                guia_resultado_multa[f'E{ultimaLinha4}'] = valorHoraMulta  # Inserindo a hora da multa
                                                guia_resultado_multa[f'F{ultimaLinha4}'] = valorCodigoMulta  # Inserindo o valor do código da multa
                                                guia_resultado_multa[f'G{ultimaLinha4}'] = valorNumeroProcessamento  # Inserindo o número de processamento
                                                guia_resultado_multa[f'H{ultimaLinha4}'] = valorSituacao  # Inserindo o valor da situação
                                                guia_resultado_multa[f'I{ultimaLinha4}'] = valorDescricaoMulta  # Inserindo o valor da descrição da multa
                                                guia_resultado_multa[f'J{ultimaLinha4}'] = valorLocalMulta  # Inserindo o valor do local da multa
                                                guia_resultado_multa[f'K{ultimaLinha4}'] = valorMunicipioMulta  # Inserindo o valor do município na multa
                                                guia_resultado_multa[f'L{ultimaLinha4}'] = valorIncluidaMulta  # Inserindo o valor de quando a multa é incluída
                                                guia_resultado_multa[f'M{ultimaLinha4}'] = valorDataLimiteDefesa  # Inserindo o valor da data limite da multa
                                                guia_resultado_multa[f'N{ultimaLinha4}'] = valorOrgaoAutuador  # Inserindo o valor do órgão autuador
                                                guia_resultado_multa[f'O{ultimaLinha4}'] = valorMulta  # Inserindo o valor da multa
                                            
                                                #Salva os resultados na planilha
                                                planilha.save(caminho_planilha)
                                                
                                                time.sleep(3)
                                            
                                            try:
                                                
                                                #Clica no botão proxima autuação
                                                botaoProxMulta = navegador.find_element(By.XPATH, '//*[@id="content"]/div[4]/div/a')
                                                navegador.execute_script("arguments[0].click();", botaoProxMulta)
                                            except:
                                                    continue
                                            
                                        #Clica para consulta a proxima autuação
                                        campoVoltarDadosVeiculos = navegador.find_element(By.CSS_SELECTOR, "#content > div:nth-child(8) > a")
                                        navegador.execute_script("arguments[0].click();", campoVoltarDadosVeiculos)
                                        
                                        time.sleep(2)
                                        
                                        #Valida a tela de erro
                                        erroSite()    
                                        
                                        #Botão de autuações
                                        campoMulta = WebDriverWait(navegador, 15).until(
                                                EC.visibility_of_element_located((By.CSS_SELECTOR, "#content > h2:nth-child(8)"))
                                        )
                                        campoMulta.click()
                                        
                                        time.sleep(2)
                                    
                                time.sleep(2)
                                guia_relacao_veiculos[f'G{linhaResult}'] = quantidadeMultas  # Inserindo a quantidade de autuações
                                planilha.save(caminho_planilha)    
                                    
                            except TimeoutException:
                                guia_relacao_veiculos[f'G{linhaResult}'] = 0  # Inserindo a quantidade de autuações
                                planilha.save(caminho_planilha)
                                print("Veículo sem multas")
                                
                            #Retorna para buscar o proximo veiculo
                            campoConsultaOutroVeiculo = WebDriverWait(navegador, 15).until(
                                EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.text-secondary[href*="/veiculos/situacao-do-veiculo/consultar-situacao-do-veiculo/"]'))
                            ).click()
                            
                            #Coloca o status na planilha
                            guia_relacao_veiculos[f'A{linhaResult}'] = "OK!"  # Inseri o status na planilha
                            planilha.save(caminho_planilha)
                            
                            
                        except Exception as e:
                            print(f"ERRO AO RESOLVER CAPTCHA")
                            
                            print(f"Tentativa {tentativas + 1}/{max_Tentativas} falhou, tentando novamente em alguns segundos...")
                            time.sleep(5)  # Espera 5 segundos antes de tentar novamente
                            
                            #Inseri a placa no campo placa
                            campoPlaca = navegador.find_element(By.CSS_SELECTOR, "#placa")
                            campoPlaca.clear()
                            campoPlaca.send_keys(placa[index])
                            
                            #Inseri o chassi no campo chassi
                            campoChassi = navegador.find_element(By.CSS_SELECTOR, "#chassi")
                            campoChassi.clear()
                            campoChassi.send_keys(chassi[index])

                            # Chama a função novamente para tentar resolver o CAPTCHA, incrementando a contagem de tentativas

                            resolver_recaptcha(tentativas + 1, max_Tentativas)
                                        

                    # Executar
                    resolver_recaptcha()
                    
                    time.sleep(2)
                        
        except TimeoutException:
            print("Reinicie o robô")

        #=======================================================================================================================

        #                                           EMITINDO OS BOLETOS


        #Navega para a pagina de emissão de multas    
        navegador.get("https://www.transito.mg.gov.br/veiculos/situacao-do-veiculo/emitir-de-extrato-de-multas")

        erroSite()

        try:
            
            campoLogin = WebDriverWait(navegador, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "#content > div > div > div.row.text-center.h5.mt-2 > div > a"))
            ).click()
        except:
            print("Prosseguir cadastro")
        #Espera que o campo placa esteja na tela 

        PaginaEmissaoMultas = WebDriverWait(navegador, 1000).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="content"]'))
        )

            
        placaEmissao = []
        renavamEmissao = []
        statusEmissao = []
        codigoProcessamentoEmissao = []
        nomeAIT = []

        ultimaLinha = 1

        for index, row in enumerate(guia_resultado_multa.iter_rows(min_row=2, max_row=guia_resultado_multa.max_row), start=0):
                
            statusEmissao.append(row[16].value) #Obtem o status
            placaEmissao.append(row[0].value) #Obtem o valor da placa
            renavamEmissao.append(row[1].value) #Obtem renavam
            codigoProcessamentoEmissao.append(row[6].value) #Obtem o codigo de processamento da multa
            nomeAIT.append(row[2].value)
            
            ultimaLinha += 1
                
            if statusEmissao[index] is None:
                    
                #Inseri a placa
                campoPlacaEmissão = navegador.find_element(By.CSS_SELECTOR, "#placa")
                campoPlacaEmissão.send_keys(placaEmissao[index])
                    
                #Inseri o renavam
                campoRenavamEmissao = navegador.find_element(By.CSS_SELECTOR, "#renavam")
                campoRenavamEmissao.send_keys(renavamEmissao[index])
                    
                #Clica no botão pesquisar
                botaoPEsquisarMUltas = navegador.find_element(By.CSS_SELECTOR, "#content > form > button")
                botaoPEsquisarMUltas.click()
                    
                erroSite()

                    #Espera o elemento de onde baixamos os boletos
                elementoPlaca = WebDriverWait(navegador, 30).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, "#content > div.row.justify-content-center.mb-3.mt-3 > div"))
                )
                    
                    
                # Localizando a tabela pela classe
                tabelaEmissao = navegador.find_element(By.CSS_SELECTOR, "#content > table")
                                
                # Contando o número de linhas (tr)
                linhasEmissao = tabelaEmissao.find_elements(By.TAG_NAME, "tr")
                numero_linhasEmissao = len(linhasEmissao)
                    
                    
                #Começa a verificar em cada uma das linha de autuação, vencidas e a vencer
                for cont in range(1, numero_linhasEmissao):
                        
                    selectorMultas = f"#content > table > tbody > tr:nth-child({cont}) > td:nth-child(1) > a"
                    cliqueEmitirMulta = navegador.find_element(By.CSS_SELECTOR, selectorMultas)
                    cliqueEmitirMulta.click()
                        
                    erroSite()
                        
                    esperaInfracoes = WebDriverWait(navegador, 10).until(
                        EC.visibility_of_element_located((By.CSS_SELECTOR, "#content > h2:nth-child(5)"))
                    )
                        
                    tabelaProcessos = navegador.find_element(By.CSS_SELECTOR, "#content > table")
                    linhaTabelaProcessos = tabelaProcessos.find_elements(By.TAG_NAME, "tr")
                    numero_LinhasTabProcessos = len(linhaTabelaProcessos)

                        
                    for cont in range(1, numero_LinhasTabProcessos):
                            
                        selectorProcesso = f"#content > table > tbody > tr:nth-child({cont}) > td:nth-child(2)"
                        campoProcesso = navegador.find_element(By.CSS_SELECTOR, selectorProcesso)
                        valorProcesso = campoProcesso.text
                        
                        time.sleep(1)
                            
                        if codigoProcessamentoEmissao[index] in valorProcesso:
                                
                            selectorImpressaoMulta = f"#content > table > tbody > tr:nth-child({cont}) > td:nth-child(6) > a > img"
                            botaoImpressao = navegador.find_element(By.CSS_SELECTOR, selectorImpressaoMulta)
                            botaoImpressao.click()
                            erroSite()
                                
                            try:
                                    
                                erroSite()
                                    
                                pyautogui.hotkey('ctrl', 's')
                                    
                                time.sleep(1)

                                caminho_arquivo = os.path.join(pasta_multas, nomeAIT[index])
                                pyautogui.write(caminho_arquivo)
                                    
                                time.sleep(2)
                                    
                                pyautogui.hotkey('enter')
                                    
                                #Coloca o status na planilha
                                guia_resultado_multa[f'Q{ultimaLinha}'] = "OK!"  # Inseri o status na planilha
                                planilha.save(caminho_planilha)

                                    
                                pyautogui.hotkey('ctrl', 'w')
                                    
                            except:
                                print("Erro ao imprimir multa")
                
                        
                    botaoVolta = navegador.find_element(By.XPATH, "//a[text()='Retornar à lista de tipos de infrações']")
                    botaoVolta.click()
                    
                    erroSite()
                    
                        #Espera o elemento de onde baixamos os boletos
                    elementoPlaca = WebDriverWait(navegador, 30).until(
                        EC.visibility_of_element_located((By.CSS_SELECTOR, "#content > div.row.justify-content-center.mb-3.mt-3 > div"))
                    )
            try:
                
                consultarOutroVeiculo = navegador.find_element(By.XPATH, '//*[@id="content"]/div[3]/a')    
                consultarOutroVeiculo.click()
            
            except NoSuchElementException:
                print("Proxima linha")
                            
        navegador.quit()

        messagebox.showinfo("Sucesso", "Automação concluída com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")

# Configuração da interface Tkinter
root = tk.Tk()
root.title("Automação DETRAN MG")
root.geometry("400x300")
root.configure(bg="#FFFFFF")  # Fundo branco

# Adicionar a logo do DETRAN-MG como ícone da janela
root.iconbitmap(r"M:\TI\ROBOS\ROBOS_EM_DEV\Automação Python\DukarDespachantes\src\PesquisaDetranMG\detran_mg.ico")

# Label inicial com texto padrão
lbl_texto = tk.Label(
    root,
    text="Selecione os caminhos entrada dos arquivos excel:",
    font=("Arial", 10, "bold"),
    justify="center",
    anchor="center",
    bg="#FFFFFF",     # Fundo branco
    fg="#4A4A4A"     # Texto cinza escuro
)
lbl_texto.pack(pady=20, padx=10, fill="both", expand=True)

# Botão para selecionar a planilha
btn_planilha = tk.Button(
    root,
    text="Caminho Planilha",
    command=selecionar_planilha,
    bg="#FFCC00",    # Amarelo vibrante
    fg="#4A4A4A",    # Texto cinza escuro
    font=("Arial", 10, "bold")
)
btn_planilha.pack(pady=10)

# Botão para iniciar a automação
btn_iniciar = tk.Button(
    root,
    text="Iniciar Automação",
    command=lambda: threading.Thread(target=iniciar_automacao).start(),
    bg="#4A4A4A",    # Cinza escuro
    fg="#FFFFFF",    # Texto branco
    font=("Arial", 10, "bold")
)
btn_iniciar.pack(pady=20)

# Créditos ao criador com meio de contato
lbl_creditos = tk.Label(
    root,
    text="Desenvolvido por Diogo Lana\nContato: diogosilvalana27@gmail.com",
    font=("Arial", 8, "italic"),
    bg="#FFFFFF",
    fg="#4A4A4A",
    anchor="center"
)
lbl_creditos.pack(side="bottom", pady=10)

# Rodar a interface
root.mainloop() 