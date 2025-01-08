#ROBO DE EMISSÃO DE IPVA E LICENCIAMENTO
from dotenv import load_dotenv
from openpyxl import load_workbook
import time
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import undetected_chromedriver as uc
import requests
from selenium.webdriver.common.keys import Keys  # Para acessar as teclas especiais
from selenium.webdriver.common.action_chains import ActionChains
import os
import pyautogui
import pdfplumber
import re

#Cria uma pasta de saída baseada no nome da planilha e em um caminho de base fornecido.
def criar_pasta_saida(caminho_planilha, pasta_downloads):

    # Pega o nome do arquivo da planilha sem a extensão
    nome_arquivo = os.path.splitext(os.path.basename(caminho_planilha))[0]

    #Caminho da pasta de saída com base na pasta fornecida
    pasta_saida = os.path.join(pasta_downloads, nome_arquivo)

    #Verifica se a pasta de saída já existe, se não existir, cria a pasta
    if not os.path.exists(pasta_saida):
        os.makedirs(pasta_saida)
        print(f"Pasta criada: {pasta_saida}")
    else:
        print(f"a pasta já existe: {pasta_saida}")

    return pasta_saida

# Caminha para selecionar o download como PDF
def selecionar_download_como_pdf(pasta_saida, placa_atual, cont):
    
    # Navegar pelas opções até "Salvar"
    for _ in range(5):  # Pressiona 'tab' 5 vezes
        pyautogui.hotkey('tab')
    time.sleep(1.5)

    pyautogui.write("salvar")  # Digita "salvar"
    time.sleep(1.5)

        # Navegar até o botão de salvar
    for _ in range(3):  # Pressiona 'tab' 6 vezes
        pyautogui.hotkey('tab')
    time.sleep(1.5)

    pyautogui.hotkey('enter')  # Confirma "Salvar"
    time.sleep(3)

    # Define o caminho completo para salvar o arquivo
    caminho_download = os.path.join(pasta_saida, f"{placa_atual}_{cont}")
    caminho_download = os.path.normpath(caminho_download)  # Normaliza o caminho

    # Digitar o caminho de salvamento
    pyautogui.write(caminho_download)
    time.sleep(1)
    pyautogui.hotkey('enter')  # Confirma o caminho

    print("Arquivo salvo com sucesso!")


pasta_downloads = r"M:\SEMINOVOS\ROBO_SEMINOVOS\EMISSAO IPVA-LIC GO"

load_dotenv()

caminho_planilha = r"C:\Users\Robo01\Desktop\ENTRADA\LOCALIZA FINAL 3 - 4 07-01.xlsx"

pasta_saida = criar_pasta_saida(caminho_planilha, pasta_downloads)

#Abri a planilha do excel
planilha = load_workbook(caminho_planilha)

#Passa a instacia da planilha BASE
guia_dados = planilha['BASE']

#Passa os cabeçalhos
guia_dados['A1'] = "PLACA"
guia_dados['B1'] = "RENAVAM"
guia_dados['D1'] = "STATUS SITE"

index = 0
linhas = list(guia_dados.iter_rows(min_row=2 , max_row= guia_dados.max_row))


# Configurar o Chrome com um User-Agent falso usando undetected-chromedriver
chrome_options = Options()
chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36")

# Inicializa o navegador com as opções e com o stealth mode ativado
navegador = uc.Chrome(options=chrome_options)
navegador.maximize_window()

navegador.get("https://www.go.gov.br/servicos/servico/consultar-veiculo--ipva-multas-e-crlv")

#Aguarda até o elemento ficar visível (com um timeout de 60 segundos)
try:
    
    #Esperar o botão de acessas e clicar nele
    botaoAcessar = WebDriverWait(navegador, 1000).until(
        EC.element_to_be_clickable((By.XPATH, "/html/body/app-root/app-main/div/div/app-sidenav/div/div/div/app-servico-detalhe/div/div[2]/div/app-botao-acessar/div/button"))
    ).click()
    
    #Esperar tela de atenção para clicar no botão entrar com o GOV
    botaoAcessarGOV = WebDriverWait(navegador, 60).until(
        EC.element_to_be_clickable((By.XPATH, "/html/body/app-root/app-modal-item[1]/div/div/div/div[2]/div/div/button"))
    ).click()
    
    time.sleep(5)
    
    campoCPF = WebDriverWait(navegador, 60).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "#accountId"))
    )
    campoCPF.send_keys(os.getenv("CPF_GOIAS"))

    botaoContinuarLogin = navegador.find_element(By.CSS_SELECTOR, "#enter-account-id")
    botaoContinuarLogin.click()

    campoSenha =  WebDriverWait(navegador, 1000).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "#password"))
    )
    campoSenha.send_keys(os.getenv("SENHA_GOIAS"))

    botaoEntrar = navegador.find_element(By.CSS_SELECTOR, "#submit-button")
    botaoEntrar.click()

    telaInicial = WebDriverWait(navegador, 60).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "exui-titulo-servico"))
    )
    
    '''
    #Validação para obter a posição na tela
    time.sleep(5)
    # Obtém a posição atual do mouse
    posicao = pyautogui.position()
    print(f"Coordenadas do mouse: {posicao}")
    '''
    
    pyautogui.leftClick(x=1197, y=394)

    time.sleep(2)
    
    botaoRealizarConsulta = WebDriverWait(navegador, 30).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "body  exui-button-secondary > button"))
    ).click()

    linhaPlan = 1
    while index < len(linhas):
        linhaPlan += 1
        row = linhas[index]

        placa_atual = row[0].value
        renavam_atual = row[1].value
        status_atual = row[2].value

        if status_atual is None:

            time.sleep(5)

            tela_pesquisa = WebDriverWait(navegador, 60).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, "exui-card-formulario-status > mat-card"))
            )

            campo_placa = navegador.find_element(By.XPATH, '(//input[contains(@class, "mat-input-element")])[1]')
            campo_placa.send_keys(placa_atual)

            time.sleep(2)

            campo_renavam = navegador.find_element(By.XPATH, '(//input[contains(@class, "mat-input-element")])[2]')
            campo_renavam.send_keys(renavam_atual)

            time.sleep(1.5)

            botao_consultar = navegador.find_element(By.CSS_SELECTOR, "exui-button-primary > button")
            botao_consultar.click()

            tela_dados_veiculo = WebDriverWait(navegador, 60).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, "body > div > div > div > div > div > lib-detalhes-veiculo > div > exui-abas > div > div > exui-aba:nth-child(1) > div > lib-dados-veiculo > div > div > exui-card-detalhamento > exui-card > mat-card > div > div"))
            )

            time.sleep(3)

            campo_debitos_veiculo = navegador.find_element(By.CSS_SELECTOR, "lib-detalhes-veiculo exui-abas > div > ul > li:nth-child(2) span")
            campo_debitos_veiculo.click()

            campo_ipva = WebDriverWait(navegador, 60).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "div:nth-child(1) > exui-card-info > div"))
            ).click()

            tabela_ipva = WebDriverWait(navegador, 60).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, "#debIpva table"))
            )

            linhas_ipva = tabela_ipva.find_elements(By.TAG_NAME, "tr")
            numero_linhas_ipva = len(linhas_ipva)


            for cont in range(1, numero_linhas_ipva):

                selector_situacao_ipva = f"#debIpva tr:nth-child({cont}) > td.mat-cell.cdk-cell.cdk-column-valorIpva.mat-column-valorIpva.ng-star-inserted"
                campo_situacao_ipva = navegador.find_element(By.CSS_SELECTOR, selector_situacao_ipva)

                situacao_ipva = campo_situacao_ipva.text 

                if situacao_ipva not in ["PAGO", "ISENTO", "QUITADO SEFAZ", "ISENTO / PAGO", "----", "PG OUTRA UF"]:

                    selector_ano_ipva = f"#debIpva tr:nth-child({cont}) > td.mat-cell.cdk-cell.cdk-column-anoExercicio.mat-column-anoExercicio.ng-star-inserted"
                    campo_ano_ipva = navegador.find_element(By.CSS_SELECTOR, selector_ano_ipva)

                    ano_ipva = campo_ano_ipva.text

                    '''
                    selector_valor_total = f"#debIpva tr:nth-child({cont}) > td.mat-cell.cdk-cell.cdk-column-valorTotal.mat-column-valorTotal.ng-star-inserted"
                    campo_valor_total = navegador.find_element(By.CSS_SELECTOR, selector_valor_total)

                    valor_total = campo_valor_total.text.strip()
                    valor_total = valor_total.replace("R$ ", "")  # Remove o prefixo "R$ "
                    valor_total = valor_total.replace(".", "")    # Remove pontos (separadores de milhar)
                    valor_total = valor_total.replace(",", ".")   # Substitui vírgula decimal por ponto
                    valor_total = float(valor_total)              # Converte para float
                    valor_total_ipvas += valor_total
                    '''
                    
                    selector_botao_pagar = f"#debIpva tr:nth-child({cont}) > td.mat-cell.cdk-cell.cdk-column-botao.mat-column-botao.ng-star-inserted > div > exui-button-primary:nth-child(1) > button"
                    botao_pagar = navegador.find_element(By.CSS_SELECTOR, selector_botao_pagar)

                    botao_pagar.click()
                    
                    tela_confirmacao = WebDriverWait(navegador, 60).until(
                        EC.visibility_of_element_located((By.CSS_SELECTOR, "#swal2-title"))
                    )
                    
                    botao_confirmar = navegador.find_element(By.CSS_SELECTOR, "body > div.swal2-container.swal2-center.swal2-backdrop-show > div > div.swal2-actions > button.swal2-confirm.exui-fill-button.swal2-styled")
                    botao_confirmar.click()
                    
                    tela_forma_pagamento = WebDriverWait(navegador, 60).until(
                        EC.visibility_of_element_located((By.XPATH, "/html/body/app-root/app-raiz-servicos-digitais/body/div/div/div/div/div/lib-detalhes-veiculo/div/exui-abas/div/div/exui-aba[2]/div/lib-debitos-veiculo/lib-modal-forma-de-pagamento/exui-modal-item/div/div/div[2]"))
                    )
                    
                    time.sleep(2)
                    
                    botao_dowload_boleto = navegador.find_element(By.XPATH, "/html/body/app-root/app-raiz-servicos-digitais/body/div/div/div/div/div/lib-detalhes-veiculo/div/exui-abas/div/div/exui-aba[2]/div/lib-debitos-veiculo/lib-modal-forma-de-pagamento/exui-modal-item/div/div/div[2]/exui-card/mat-card/mat-card-content/div/div[2]/div[2]/exui-button-secondary/button")
                    botao_dowload_boleto.click()
                    
                    tela_downloads = WebDriverWait(navegador, 60).until(
                        EC.visibility_of_element_located((By.CSS_SELECTOR, "div.header-modal-wrap"))
                    )
                    
                    '''
                    #Validação para obter a posição na tela
                    time.sleep(5)
                    # Obtém a posição atual do mouse
                    posicao = pyautogui.position()
                    print(f"Coordenadas do mouse: {posicao}")
                    '''    
                    
                    time.sleep(5)
                    
                    # Coordenadas onde o clique será feito
                    x, y = 1113, 188

                    # Clique com o botão direito na posição especificada
                    pyautogui.leftClick(x=x, y=y)
                    
                    time.sleep(2)
                    
                    selecionar_download_como_pdf(pasta_saida, placa_atual, cont)
                    
                    time.sleep(2)
                    
                    '''
                    #Validação para obter a posição na tela
                    time.sleep(5)
                    # Obtém a posição atual do mouse
                    posicao = pyautogui.position()
                    print(f"Coordenadas do mouse: {posicao}")
                    '''
                    pyautogui.leftClick(x=1141, y=137)
                    
            guia_dados[f'C{linhaPlan}'] = "OK!"
            

            botao_voltar = WebDriverWait(navegador, 60).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "lib-detalhes-veiculo > div > div > exui-button-no-border > button"))
            ).click()
            
        index += 1
        planilha.save(caminho_planilha)
                    
except TimeoutException:
    
    
    print("Renicie a automação")
    breakpoint()
    
navegador.quit()

#TRATATIVA BOLETOS

arquivos_pdfs = [f for f in os.listdir(pasta_saida) if f.endswith('.pdf')]

for arquivos in arquivos_pdfs:
    
    caminho_pdf = os.path.join(pasta_saida, arquivos)
    
    with pdfplumber.open(caminho_pdf) as pdf:
        for num_page, pagina in enumerate(pdf.pages):
            texto_boleto = pagina.extract_text()
        
            # Extrair valor total
            corte_valor_total1 = re.split("Sacador/Avalista CPF/CNPJ ", texto_boleto)
            corte_valor_total2 = re.split("\nAutenticação Mecânica", corte_valor_total1[1])
            valor_total = corte_valor_total2[0].strip()
            valor_total = valor_total.replace('.', '').replace(',', '.')  # Ajustar o formato do número
            
            # Extrair proprietário
            corte_propietarario1 = re.split("PROPRIETÁRIO: ", texto_boleto)
            corte_propietarario2 = re.split("\nCPF/CNPJ:", corte_propietarario1[1])
            valor_propietario = corte_propietarario2[0].strip()
            
            # Extrair valor do licenciamento
            padrao = r"LICENCIAMENTO ANUAL \[2025\]\s+([\d.,]+)"
            resultado = re.search(padrao, texto_boleto)
            valor_licenciamento = resultado.group(1).replace('.', '').replace(',', '.')  # Ajustar o formato do número
            
            # Calcular valor do IPVA
            valor1 = float(valor_total)
            valor2 = float(valor_licenciamento)
            valor_ipva = valor1 - valor2
            
            # Extrair a placa
            padrao_placa = r"PLACA:\s+([A-Z0-9]+)"
            resultado = re.search(padrao_placa, texto_boleto)
            placa = resultado.group(1)
            
            # Exibir os valores extraídos
            print(f"Arquivo: {placa}")
            print(f"Proprietário: {valor_propietario}")
            print(f"Valor Total: {valor1:.2f}")
            print(f"Licenciamento: {valor2:.2f}")
            print(f"IPVA: {valor_ipva:.2f}")
            print("-" * 40)
            
            # Procurar a placa na planilha e preencher dados
            for linha in range(2, guia_dados.max_row + 1):
                celula_placa = guia_dados[f'A{linha}'].value  # Coluna 'A' contém as placas
                if celula_placa == placa:  # Verificar se a placa da planilha corresponde à do boleto
                    
                    guia_dados[f'G{linha}'] = valor_propietario  # Coluna 'B': PROPRIETARIO
                    guia_dados[f'F{linha}'] = valor_total  # Coluna 'C': VALOR TOTAL
                    guia_dados[f'E{linha}'] = valor_licenciamento  # Coluna 'D': VALOR LIC
                    guia_dados[f'D{linha}'] = valor_ipva  # Coluna 'E': VALOR IPVA
                    guia_dados[f'H{linha}'] = "BOLETO PROCESSADO"  # Coluna 'F': STATUS BOLETO
                    print(f"Dados preenchidos para a placa {placa} na linha {linha}.")
                     # Salvar a planilha
                    planilha.save(caminho_planilha)
                    break
            else:
                print(f"Placa {placa} não encontrada na planilha.")
    
