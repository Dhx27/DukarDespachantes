#ROBO EMISSÃO DE DOCUMENTOS
from openpyxl import load_workbook
import undetected_chromedriver as uc
import time
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from dotenv import load_dotenv

CnpjCliente = "07563781000171"
pathEntrada = r'C:\Users\diogo.lana\Desktop\TESTE\MODELO (2).xlsx'
pathSaida = r''

# Configurar o Chrome com um User-Agent falso usando undetected-chromedriver
chrome_options = Options()
chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36")

# Inicializa o navegador com as opções e com o stealth mode ativado
navegador = uc.Chrome(options=chrome_options)
navegador.maximize_window()

time.sleep(5)

# Acessa o site especificado
navegador.get("https://portalservicos.senatran.serpro.gov.br/#/login")

# Aguarda até o elemento ficar visível (com um timeout de 60 segundos)
try:
    telaAcesso = WebDriverWait(navegador, 60).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "body > modal-container > div > div > div.modal-header > h4"))
    )
except TimeoutException:
    print("Tela de acesso não apareceu a tempo.")
    navegador.refresh()

time.sleep(5)

# Clica em "Representante Jurídico"
representanteJuridico = navegador.find_element(By.CSS_SELECTOR, ".modal-content ul > li:nth-child(2) a")
representanteJuridico.click()

# Aguarda até a tela de CNPJ ficar visível
try:
    telaCnpj = WebDriverWait(navegador, 30).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "body > modal-container > div > div > div.modal-body.view-dashboard > div:nth-child(4) > div:nth-child(1) > div > h5"))
    )
except TimeoutException:
    print("Tela de CNPJ não apareceu.")

time.sleep(5)

# Seleciona o CNPJ no dropdown
select_cnpj = navegador.find_element(By.CSS_SELECTOR, "body > modal-container > div > div > div.modal-body.view-dashboard > div:nth-child(4) > div:nth-child(2) > div > div > select")
select = Select(select_cnpj)
select.select_by_value(CnpjCliente)

time.sleep(3)

# Clicar após selecionar o cliente
botaoSelecionarCleinte = navegador.find_element(By.CSS_SELECTOR, "body > modal-container > div > div > div.modal-body.view-dashboard > div.modal-footer > div > button.br-button.primary.small.footer-button")
botaoSelecionarCleinte.click()

time.sleep(5)

# Espera a tela de consulta de veículos
try:
    telaVeiculos = WebDriverWait(navegador, 30).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "div.view-dashboard li:nth-child(1)"))
    )
except TimeoutException:
    print("Botão de veículos não estava visível.")

# Tenta fechar qualquer aba indesejada
try:
    fecharAba = navegador.find_element(By.CSS_SELECTOR, "body > modal-container > div > div > div.modal-footer.alt-footer.text-center > div > button.br-button.secondary.small.footer-button")
    fecharAba.click()
except NoSuchElementException:
    print("Botão de fechar aba não encontrado.")

# Aceita cookies
cookies = navegador.find_element(By.CSS_SELECTOR, "#cookiebar > div.br-cookiebar.default > div > div > div > div.br-modal-footer.actions.justify-content-end > button.br-button.primary.small")
cookies.click()

time.sleep(3)

# Clica na opção de consulta de veículos
try:
    botaoVeiculos = navegador.find_element(By.CSS_SELECTOR, "body > app-root > form > br-main-layout > div > div > main > app-usuario > app-home > div > div.view-dashboard > ul > li:nth-child(1) > a")
    botaoVeiculos.click()
except NoSuchElementException:
    print("Botão de veículos não encontrado.")

time.sleep(5)

#-------------------------------------------------------------------------------

# Função para aguardar o elemento de pesquisa no campo "PLACA"
def esperarCampoConsultaVeiculos(navegador, timeout=30):
    campoCSS = (
        "body > app-root > form > br-main-layout > div > div > main > app-veiculo > app-veiculos-list > div > div > div > form > br-tab-set > div > nav > br-tab > div > div:nth-child(2) > div:nth-child(1) > br-input > div > div > input"
    )
    
    try:
        campoConsultaVeiculos = WebDriverWait(navegador, timeout).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, campoCSS))
        )
        return campoConsultaVeiculos  # Retorna o elemento se for encontrado
    except TimeoutException:
        print("Campo consulta não encontrado")
        return None  # Retorna None se o elemento não foi encontrado


# Usando a Função (esperarCampoConsultaVeiculos) para aguardar o campo
consultaVeiculos = esperarCampoConsultaVeiculos(navegador)

# Se o campo não aparecer recarrega a página
if not consultaVeiculos:
    navegador.refresh()

# Recarrega a página se o senatran apresentar algum erro
try:
    erroValidacaoCampo = navegador.find_element(By.CSS_SELECTOR, "body > app-root > form > br-main-layout > div > div > main > br-alert-messages > div > div > div.content > div")
    if erroValidacaoCampo.text == "Ocorreu erro na validação do campo abaixo:":
        navegador.refresh()
except NoSuchElementException:
    print("Erro não encontrado")

# Carrega a planilha
worbook = load_workbook(pathEntrada)
instanciaExcel = worbook.active

# Itera sobre as linhas da planilha
for linhaDaVez in range(2, instanciaExcel.max_row + 1):  # Assume que a primeira linha é o cabeçalho
    placa = instanciaExcel[f'A{linhaDaVez}'].value  # Obtém a placa
    status = instanciaExcel[f'B{linhaDaVez}'].value  # Obtém o status
    
    if status != "OK!":
        
        tentativa_sucesso = False
        tentativas = 0
        max_tentativas = 3  # Defina quantas vezes deseja tentar
        
        while not tentativa_sucesso and tentativas < max_tentativas:
        
            print(f"Pesquisando placa da vez: {placa}")
            
            #Confere se o campo para baixar o documento esta visivel
             
            try:
                
                #Seleciona o campo placa
                campoPlaca = WebDriverWait(navegador, 20).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "body > app-root > form > br-main-layout > div > div > main > app-veiculo > app-veiculos-list > div > div > div > form > br-tab-set > div > nav > br-tab > div > div:nth-child(2) > div:nth-child(1) > br-input > div > div > input"))
                )
                campoPlaca.clear() # Limpa o campo placa
                campoPlaca.send_keys(placa) # Digita a placa da vez, no campo selecionado
                
                time.sleep(3)
                
                #Seleciona e clica no botão filtrar 
                botaoFiltrar = navegador.find_element(By.CSS_SELECTOR, "body > app-root > form > br-main-layout > div > div > main > app-veiculo > app-veiculos-list > div > div > div > form > br-tab-set > div > nav > br-tab > div > div:nth-child(2) > div.col-md-6 > button.br-button.small.primary.side-button")
                navegador.execute_script("arguments[0].click();", botaoFiltrar)
            
           
                campoVeiculoDaVez = WebDriverWait(navegador, 30).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, "body > app-root > form > br-main-layout > div > div > main > app-veiculo > app-veiculos-list > div > div > div > form > br-tab-set > div > nav > br-tab > div > div.ng-star-inserted > div > div"))
                )
                
                time.sleep(2)
                
                #Clica no campo para consultar a placa do veículo   
                campoVeiculoDaVez.click()
                
                #Espera a página onde está localizado o documento
                telaBaixarDocumento = WebDriverWait(navegador, 30).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, "#header-small > br-tab-set > div > nav > br-tab:nth-child(2) > div > app-veiculo-dados > div > div > table > tbody > tr:nth-child(1) > th"))
                )
                
                #Aguarda que o botão de download esteja clicavel
                botaoBaixar = WebDriverWait(navegador, 30).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "#header-small > br-tab-set > div > nav > br-tab:nth-child(2) > div > app-veiculo-dados > div > div > table > tbody > tr:nth-child(2) > td > a"))
                )
                # Executa o clique usando JavaScript
                navegador.execute_script("arguments[0].click();", botaoBaixar)
                
                time.sleep(2)
                
                try:
                    erro = navegador.find_element(By.CSS_SELECTOR, "body > app-root > form > br-main-layout > div > div > main > br-alert-messages > div > div > div.content > div")
                        
                        # Verifica o texto do erro
                    if "Ocorreu erro na validação" in erro.text:
                            navegador.refresh()
                            botaoBaixar = navegador.find_element(By.CSS_SELECTOR, "#header-small > br-tab-set > div > nav > br-tab:nth-child(2) > div > app-veiculo-dados > div > div > table > tbody > tr:nth-child(2) > td > a")
                            navegador.execute_script("arguments[0].click();", botaoBaixar)
                            
                except NoSuchElementException:  # Captura o erro específico quando o elemento não é encontrado
                    print("Processando download")
                    
                #Escreva na linha da vez o novo status
                instanciaExcel[f'B{linhaDaVez}'] = "OK!"
                
                #Salva as alterações na planilha
                worbook.save(pathEntrada)
                
                tentativa_sucesso = True  # Marca a tentativa como bem-sucedida
                
                time.sleep(5)
                    
            except TimeoutException:
                tentativas += 1
                navegador.get("https://portalservicos.senatran.serpro.gov.br/#/veiculos/meus-veiculos")
                time.sleep(5)
                if tentativas == max_tentativas:
                    instanciaExcel[f'B{linhaDaVez}'] = "Erro ao pesquisar cliente"
                    
                
        #Retorna na página incial para continuar a pesquisa
        navegador.get("https://portalservicos.senatran.serpro.gov.br/#/veiculos/meus-veiculos")
        
        # Usando a Função (esperarCampoConsultaVeiculos) para aguardar o campo
        consultaVeiculos = esperarCampoConsultaVeiculos(navegador)
        
        
        # Se o campo não aparecer recarrega a página
        if not consultaVeiculos:
             navegador.refresh()
             
        if(placa is None):
            print("Pesquisa Finalizada!")
            # Fechar o navegador
            navegador.quit()
        