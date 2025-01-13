#ROBO DE EMISSÃO DE LICENCIAMENTO E IPVA

from dotenv import load_dotenv
from openpyxl import load_workbook
import time
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import undetected_chromedriver as uc
import requests
from selenium.webdriver.common.keys import Keys  # Para acessar as teclas especiais
from selenium.webdriver.common.action_chains import ActionChains
import os
import pyautogui
import pdfplumber
import re
import pyautogui
import base64
import pdfplumber
import re

load_dotenv()

pasta_download = r'C:\Users\Diogo Lana\Desktop\Nova pasta'
caminho_planilha = r'C:\Users\Diogo Lana\Desktop\Nova pasta\MODELO MT.xlsx'

# Dados do Turnstile CAPTCHA
API_KEY = os.getenv("api_key")
SITEKEY = "0x4AAAAAAAO9omZCUc8pnQfN"
PAGE_URL = "https://internet.detrannet.mt.gov.br/ConsultaVeiculo.asp"

def selecionar_download_como_pdf_ipva(pasta_saida, placa_atual_ipva):
    time.sleep(5)
    
    pyautogui.hotkey('ctrl', 'p')
    time.sleep(3)
    
    # Navegar pelas opções até "Salvar"
    for _ in range(5):  # Pressiona 'tab' 5 vezes
        pyautogui.hotkey('tab')
    time.sleep(1.5)

    pyautogui.write("salvar")  # Digita "salvar"
    time.sleep(1.5)

    # Navegar até o botão de salvar
    for _ in range(3):  
        pyautogui.hotkey('tab')
    time.sleep(1.5)

    pyautogui.hotkey('enter')  # Confirma "Salvar"
    time.sleep(3)

    # Define o caminho completo para salvar o arquivo
    caminho_download = os.path.join(pasta_saida, f"IPVA {placa_atual_ipva}")
    caminho_download = os.path.normpath(caminho_download)  # Normaliza o caminho

    # Digitar o caminho de salvamento
    pyautogui.write(caminho_download)
    time.sleep(1)
    pyautogui.hotkey('enter')  # Confirma o caminho

    print("Arquivo salvo com sucesso!")

def selecionar_download_como_pdf_lic(pasta_saida, placa_atual):
    
    time.sleep(15)
    
    pyautogui.hotkey('ctrl', 'p')
    time.sleep(3)
    
    # Navegar pelas opções até "Salvar"
    for _ in range(5):  # Pressiona 'tab' 5 vezes
        pyautogui.hotkey('tab')
    time.sleep(1.5)

    pyautogui.write("salvar")  # Digita "salvar"
    time.sleep(1.5)

        # Navegar até o botão de salvar
    for _ in range(4):  # Pressiona 'tab' 6 vezes
        pyautogui.hotkey('tab')
    time.sleep(1.5)

    pyautogui.hotkey('enter')  # Confirma "Salvar"
    time.sleep(3)

    # Define o caminho completo para salvar o arquivo
    caminho_download = os.path.join(pasta_saida, f"LIC {placa_atual}")
    caminho_download = os.path.normpath(caminho_download)  # Normaliza o caminho

    # Digitar o caminho de salvamento
    pyautogui.write(caminho_download)
    time.sleep(1)
    pyautogui.hotkey('enter')  # Confirma o caminho

    print("Arquivo salvo com sucesso!")
    time.sleep(1)
    pyautogui.hotkey('ctrl', 'w')
    time.sleep(1.5)
    pyautogui.hotkey('ctrl', 'w')
    

# 1. Enviar requisição para resolver o CAPTCHA turnstile
def enviar_requisicao_captcha(api_key, sitekey, page_url):
    url = "http://2captcha.com/in.php"
    data = {
        "key": api_key,
        "method": "turnstile",
        "sitekey": sitekey,
        "pageurl": page_url,
    }
    response = requests.post(url, data=data)
    if response.status_code == 200 and "OK|" in response.text:
        captcha_id = response.text.split('|')[1]
        return captcha_id
    else:
        raise Exception(f"Erro ao enviar captcha: {response.text}")

# 2. Obter o token de resposta do CAPTCHA
def obter_resposta_captcha(api_key, captcha_id):
    url = f"http://2captcha.com/res.php?key={api_key}&action=get&id={captcha_id}"
    while True:
        response = requests.get(url)
        if "CAPCHA_NOT_READY" in response.text:
            print("Captcha ainda não resolvido, tentando novamente em 5 segundos...")
            time.sleep(5)
        elif "OK|" in response.text:
            return response.text.split('|')[1]
        else:
            raise Exception(f"Erro ao obter resposta do captcha: {response.text}")

# 3. Inserir o token na página (use Selenium ou outro método)
def inserir_token(navegador, token):
    try:

        elemento = WebDriverWait(navegador, 30).until(
            EC.presence_of_element_located((By.XPATH, "//input[@name='g-recaptcha-response']"))
        )
        
        valor_captcha = f"{token}"
        navegador.execute_script(f"arguments[0].setAttribute('value', '{valor_captcha}')", elemento)

        # Disparar o evento de mudança (opcional)
        #navegador.execute_script("arguments[0].dispatchEvent(new Event('change'))", elemento)

    except Exception as e:
        raise Exception(f"Erro ao inserir token: {str(e)}")



#Cria uma pasta de saída baseada no nome da planilha e em um caminho de base fornecido.
def criar_pasta_saida(pasta_download, caminho_planilha):

    #Pega o nome do arquivo excel sem a extensão
    nome_arquivo = os.path.splitext(os.path.basename(caminho_planilha))[0]

    #Caminho da pasta saida com base na pasta fornecida
    pasta_saida = os.path.join(pasta_download, nome_arquivo)

    #Verifica se a pasta de saída já existe, se não existir, cria a pasta
    if not os.path.exists(pasta_saida):
        os.makedirs(pasta_saida)
        print(f"Pasta criada: {pasta_saida}")
    else:
        print(f"A pasta já está criada {pasta_saida}")

    return pasta_saida

'''
# Configurar o Chrome com um User-Agent falso usando undetected-chromedriver
chrome_options = Options()
chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36")

# Inicializa o navegador com as opções e com o stealth mode ativado
navegador = uc.Chrome(options=chrome_options)
navegador.maximize_window()
'''
pasta_saida = criar_pasta_saida(pasta_download, caminho_planilha)

#Abri a planilha do excel
planilha = load_workbook(caminho_planilha)

#Passa a instacia da planilha BASE
guia_dados = planilha['BASE']

"""
#                           EMISSÃO LICENCIAMENTO 

navegador.get("https://www.detran.mt.gov.br/")

try:
    
    tela_atendimento = WebDriverWait(navegador, 10).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "#myPopup > a > picture > img"))
    )
            
    botao_fechar = navegador.find_element(By.CSS_SELECTOR, "#myPopup > span")
    navegador.execute_script("arguments[0].click();", botao_fechar)

except (TimeoutException, NoSuchElementException):  
     pass
 
tela_cookies = WebDriverWait(navegador, 30).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="adopt-accept-all-button"]'))
).click()
        

# Pegar todas as guias abertas
abas = navegador.window_handles


index = 0
linhas = list(guia_dados.iter_rows(min_row=2, max_row=guia_dados.max_row))
linhaPlan = 1


try:

    while index < len(linhas):
        linhaPlan += 1

        row = linhas[index]

        #Obtem todos os valores dispobniveis na planilha
        placa_atual = row[0].value
        renavam_atual = row[1].value
        cnpj_atual = row[2].value
        status_lic = row[3].value

        if status_lic is None:
        
            time.sleep(3)
            
            campo_placa_renavam = WebDriverWait(navegador, 100).until(
                EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div[4]/section/div/div/div/div/div[1]/section/div/div/div/div/div/div[2]/div/div/section/div/div[2]/div/div/div/div/div'))
            )
            
            campo_placa = navegador.find_element(By.CSS_SELECTOR, "#input_placa")
            campo_placa.clear()
            campo_placa.send_keys(placa_atual)
            
            time.sleep(1.5)
            
            campo_renavam = navegador.find_element(By.CSS_SELECTOR, "#input_renavam")
            campo_renavam.clear()
            campo_renavam.send_keys(renavam_atual)
            
            botao_consultar = WebDriverWait(navegador, 30).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "#formVeiculo > div:nth-child(4) > input.dtrn-frm-sub.dtrn-vin.text-size-acessibilidade"))
            ).click()
            
            # Pegar todas as guias abertas
            abas = navegador.window_handles
            
            # Alternar para a nova guia (última aberta)
            navegador.switch_to.window(abas[-1])
            print("Estamos na nova guia:", navegador.title)
            
            tela_cnpj = WebDriverWait(navegador, 60).until(
                EC.visibility_of_element_located((By.XPATH, '/html/body/center/div/form/table/tbody/tr[3]/td[2]/input'))
            )

            time.sleep(5)
            
            try:
                print("Enviando requisição para resolver o CAPTCHA...")
                captcha_id = enviar_requisicao_captcha(API_KEY, SITEKEY, PAGE_URL)
                print("Obtendo resposta do CAPTCHA...")
                token = obter_resposta_captcha(API_KEY, captcha_id)
                print("Solução do CAPTCHA obtida:")
                inserir_token(navegador, token)  # Descomente ao integrar com Selenium
                
                cnpj = cnpj_atual.replace(".", "").replace("/", "").replace("-", "")
                
                campo_cnpj = navegador.find_element(By.XPATH, '/html/body/center/div/form/table/tbody/tr[3]/td[2]/input')
                campo_cnpj.send_keys(cnpj)
                
                botao_consultar_veiculo = WebDriverWait(navegador, 60).until(
                    EC.element_to_be_clickable((By.XPATH, '/html/body/center/div/form/table/tbody/tr[4]/td/input'))
                ).click()

                try:

                    campo_placa_renavam_nao_comferem = WebDriverWait(navegador, 10).until(
                        EC.visibility_of_element_located((By.CSS_SELECTOR, "body > center > div > table > tbody > tr > td"))
                    )
                    
                    elemento_placa_renavam_nao_comferem = campo_placa_renavam_nao_comferem.text

                    if "Documento Proprietario informado não confere com o proprietario do veiculo" in elemento_placa_renavam_nao_comferem:
                        guia_dados[f'D{linhaPlan}'] = "Documento Proprietario informado não confere com o proprietario do veiculo."

                    if "Placa e renavam não conferem" in elemento_placa_renavam_nao_comferem:
                        guia_dados[f'D{linhaPlan}'] = "Placa e renavam não conferem"

                    guia_dados[f'D{linhaPlan}'] = "Placa e renavam não conferem"
                    planilha.save(caminho_planilha)
                    index += 1
                    
                    navegador.close()
                    
                    # Voltar para a guia original
                    navegador.switch_to.window(abas[0])
                    print("Voltamos para a guia original:", navegador.title)
                    continue

                except (NoSuchElementException, TimeoutException):
                    pass
                
                tela_baixar_boleto = WebDriverWait(navegador, 60).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, "#exibir_servicos_Debitos"))
                )
                
                #modal_licenciamento = navegador.find_element(By.CSS_SELECTOR, "#cmbTipoDebito")
                #modal_licenciamento.click()
                
                elemento_tabela = navegador.find_element(By.CSS_SELECTOR, "#div_servicos_Debitos > table > tbody > tr > td")
                campo_debitos = elemento_tabela.text
                
                if campo_debitos == "Nenhum débito em aberto cadastrado para este veículo.":
                    guia_dados[f'D{linhaPlan}'] = campo_debitos
                    planilha.save(caminho_planilha)
                    index += 1
                    
                    navegador.close()
                    
                    # Voltar para a guia original
                    navegador.switch_to.window(abas[0])
                    print("Voltamos para a guia original:", navegador.title)
                    continue

                    
                
                botao_download_guia = WebDriverWait(navegador, 60).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "#spanDAR_LicenciamentoExercicio"))
                ).click()
                    
                time.sleep(5)
                
                selecionar_download_como_pdf_lic(pasta_saida, placa_atual)
                
                guia_dados[f'D{linhaPlan}'] = "OK!"  
                planilha.save(caminho_planilha)
                
                # Voltar para a guia original
                navegador.switch_to.window(abas[0])
                print("Voltamos para a guia original:", navegador.title)
                
                index += 1
                
            except Exception as e:
                print(f"Erro: {str(e)}")
        else:
            index += 1
            continue

except TimeoutException:
    print("Renicie a automação")
    breakpoint()



#                                       EMISSÃO IPVA   

#Enviar requisição para resolver o captcha do ipva
def enviar_requisicao_captcha_1(api_key, base64_image):
    
    url = "http://2captcha.com/in.php"
    data = {
        "key": api_key,
        "method": "base64",
        "body": base64_image,
        "json": 1,
    }

    response = requests.post(url, data=data).json()
    if response["status"] == 1:  # Status 1 indica sucesso
        captcha_id = response["request"]
        return captcha_id
    else:
        raise Exception(f"Erro ao enviar captcha: {response}")
    
# 2. Obter o token de resposta do CAPTCHA
def obter_resposta_captcha_1(api_key, captcha_id):
    
    url = f"http://2captcha.com/res.php?key={api_key}&action=get&id={captcha_id}&json=1"
    while True:
        
        response = requests.get(url).json()
        if response["status"] == 1:
            return response["request"]
        elif response["request"] == "CAPCHA_NOT_READY":
            print("Captcha ainda não resolvido, aguardando...")
        else:
            raise Exception(f"Erro ao obter resposta: {response}")
        time.sleep(5)
 
 
def quebra_captcha():
    
    try:
        # Esperar e localizar a imagem do CAPTCHA
        imagem_captcha = WebDriverWait(navegador, 10).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "body > div > p:nth-child(2) > img:nth-child(1)"))
        )

        # Obter a URL base64 da imagem do CAPTCHA
        base64_url = imagem_captcha.get_attribute("src")
        base64_data = base64_url.split(",")[1]

        # Enviar CAPTCHA para o 2Captcha
        captcha_id = enviar_requisicao_captcha_1(API_KEY, base64_data)

        # Obter a resposta do CAPTCHA
        resposta_captcha = obter_resposta_captcha_1(API_KEY, captcha_id)
        print(f"Resposta do CAPTCHA: {resposta_captcha}")

        # Inserir a resposta no campo do CAPTCHA
        campo_captcha = WebDriverWait(navegador, 10).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "#ans"))
        )
        campo_captcha.clear()
        campo_captcha.send_keys(resposta_captcha)

        # Submeter o formulário
        botao_confirmar = WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "#jar"))
        )
        botao_confirmar.click()

    except (TimeoutException, NoSuchElementException) as e:
        pass
        
        
navegador.get("https://www.sefaz.mt.gov.br/ipva/emissaoguia/emitir")

quebra_captcha()

index = 0
linhas = list(guia_dados.iter_rows(min_row=2, max_row=guia_dados.max_row))
linhaPlan = 1

try:
    while index < len(linhas):
        linhaPlan += 1

        row = linhas[index]
        
        placa_atual_ipva = row[0].value
        renavam_atual_ipva = row[1].value
        status_atual_ipva = row[4].value
        
        if status_atual_ipva is None:
            
            tela_pesquisa_ipva = WebDriverWait(navegador, 60).until(
                EC.visibility_of_element_located((By.XPATH, '/html/body/center/form/table/tbody/tr[5]/td[2]/input'))
            )
            
            campo_input_renavam = navegador.find_element(By.XPATH, '/html/body/center/form/table/tbody/tr[5]/td[2]/input')
            campo_input_renavam.clear()
            campo_input_renavam.send_keys(renavam_atual_ipva)
            
            botao_consultar = WebDriverWait(navegador, 60).until(
                EC.element_to_be_clickable((By.XPATH, '/html/body/center/form/p/input[1]'))
            ).click()
            
            #Aqui pedi captch
            quebra_captcha()
            
            tela_ipva = WebDriverWait(navegador, 60).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, "table.SEFAZ-TABLE-Moldura"))
            )
            
            elemento_desconto5 = WebDriverWait(navegador, 60).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "#tipoGuia\.2025\.2"))
            ).click()
            
            try:
                
                campo_email = WebDriverWait(navegador, 10).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, "#email"))
                ).send_keys(os.getenv("email_dukar"))
                
                campo_numero = WebDriverWait(navegador, 10).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, "#celular"))
                ).send_keys(os.getenv("numero_dukar"))
                
            except (NoSuchElementException, TimeoutException):
                pass
            
            botao_gerar_guias = WebDriverWait(navegador, 60).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "#btnOK"))
            ).click()
            
            try:
                quebra_captcha() 
                
            except(NoSuchElementException, TimeoutException):
                pass
            
            quebra_captcha()
                        
            time.sleep(5)
            
            selecionar_download_como_pdf_ipva(pasta_saida, placa_atual_ipva)
            
            guia_dados[f'E{linhaPlan}'] = "OK!"  
            planilha.save(caminho_planilha)
            
            index +=1
            
            navegador.get("https://www.sefaz.mt.gov.br/ipva/emissaoguia/emitir")            
            
        else:
            
            index += 1 
            continue
        
except TimeoutException:
    print("Renicie a automação")
    breakpoint()

navegador.quit()
"""
#                       MANIPULAÇÃO DE PDFS

arquivos_pdfs = [f for f in os.listdir(pasta_saida) if f.endswith('.pdf')]

for arquivos in arquivos_pdfs:
    
    caminho_pdf = os.path.join(pasta_saida, arquivos)
    
    with pdfplumber.open(caminho_pdf) as pdf:
        for num_page, pagina in enumerate(pdf.pages):
            
            texto_boleto = pagina.extract_text()
            
            if "Licenciamento" in texto_boleto:
                
                #Extrai a placa do boleto 
                corte_placa_1 = re.split("Placa:", texto_boleto)
                corte_placa_2 = re.split(" / ", corte_placa_1[1])
                placa = corte_placa_2 [0]
                
                #Extrai valor do licenciamento do boleto
                corte_valor_licenciamento_1 = re.split("TOTAL A RECOLHER 31 - VALOR", texto_boleto)
                corte_valor_licenciamento_2 = re.split("33 - VALOR A RECOLHER POR EXTENSO", corte_valor_licenciamento_1[1])
                valor_licenciamento = corte_valor_licenciamento_2[0].replace("\n", "")
                
                corte_propietario_1 = re.split("CNPJ OU CPF SELO FISCAL NA SAÍDA", texto_boleto)
                corte_propietario_2 = re.split("PARA OUTRA U.F.", corte_propietario_1[1])
                propietario = corte_propietario_2[0].replace("\n", "")
            
            if "IPVA" in texto_boleto:
                
                #Extrai a placa do boleto 
                corte_placa_1 = re.split("Placa : ", texto_boleto)
                corte_placa_2 = re.split("JUROS 29 - VALOR", corte_placa_1[1])
                placa = corte_placa_2 [0].replace(" ", "")
                
                corte_valor_ipva_1 = re.split("3. Escaneie o código abaixo ", texto_boleto)
                corte_valor_ipva_2 = re.split("4. Confira as informações", corte_valor_ipva_1[1])
                valor_ipva = corte_valor_ipva_2[0].replace("\n", "")
                
                # Procurar a placa na planilha e preencher dados
                for linha in range(2, guia_dados.max_row + 1):
                    celula_placa = guia_dados[f'A{linha}'].value  # Coluna 'A' contém as placas
                    if celula_placa == placa: 
                        guia_dados[f'F{linha}'] = valor_ipva  # Coluna 'E': VALOR IPVA
                        planilha.save(caminho_planilha)
                        break
                    else:
                        print(f"Placa {placa} não encontrada na planilha.")
                
                continue
           
            # Procurar a placa na planilha e preencher dados
            for linha in range(2, guia_dados.max_row + 1):
                celula_placa = guia_dados[f'A{linha}'].value  # Coluna 'A' contém as placas
                if celula_placa == placa:  # Verificar se a placa da planilha corresponde à do boleto
                    guia_dados[f'I{linha}'] = propietario  # Coluna 'B': PROPRIETARIO
                    guia_dados[f'G{linha}'] = valor_licenciamento  # Coluna 'D': VALOR LIC
                    guia_dados[f'J{linha}'] = "BOLETO PROCESSADO"  # Coluna 'F': STATUS BOLETO
                    print(f"Dados preenchidos para a placa {placa} na linha {linha}.")
                     # Salvar a planilha
                    planilha.save(caminho_planilha)
                    break
            else:
                print(f"Placa {placa} não encontrada na planilha.")

for linha in range (2, guia_dados.max_row + 1):
    
    celula_ipva = guia_dados[f'F{linha}'].value
    celula_licenciamento = guia_dados[f'G{linha}'].value
    
    # Verifica se algum dos valores é None ou está vazio
    if not celula_ipva or not celula_licenciamento:
        continue  # Pula para a próxima iteração se qualquer valor estiver vazio
    
    valor_ipva_calc = float(celula_ipva.replace('.', '').replace(',', '.'))
    valor_licenciamento_cal = float(celula_licenciamento.replace('.', '').replace(',', '.'))
    
    valor_total = valor_ipva_calc + valor_licenciamento_cal
    
    guia_dados[f'H{linha}'] = valor_total  # Coluna 'H': VALOR 
    planilha.save(caminho_planilha)
    
