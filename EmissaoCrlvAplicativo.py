import os
import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from twocaptcha import TwoCaptcha
import pyautogui
import tkinter as tk
from tkinter import filedialog, messagebox
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager  # Importa o WebDriver Manager
from selenium.webdriver.chrome.service import Service  # Importar Service


# Função para selecionar o caminho da planilha
def selecionar_planilha():
    caminho = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if caminho:
        global caminhoPlanilha
        caminhoPlanilha = caminho
        label_planilha.config(text=f"Caminho da Planilha: {os.path.basename(caminhoPlanilha)}")

# Função para selecionar o caminho dos processados
def selecionar_processados():
    caminho = filedialog.askdirectory()
    if caminho:
        global pathDownloads
        pathDownloads = caminho
        label_processados.config(text=f"Caminho dos Processados: {os.path.basename(pathDownloads)}")

# Função principal que executa o código de automação
def executar_automacao():
    if not caminhoPlanilha or not pathDownloads:
        messagebox.showerror("Erro", "Por favor selecione ambos os caminhos.")
        return

    api_key = os.getenv('API_KEY_2CAPTCHA', 'CHAVE')  # Use sua própria chave API
    solver = TwoCaptcha(api_key)

    service = Service(ChromeDriverManager().install())  # Cria o serviço do ChromeDriver
    navegador = webdriver.Chrome(service=service)  # Passa o serviço na inicialização do webdriver
    navegador.maximize_window()
    
    caminho = os.path.normpath(pathDownloads)

    try:
        navegador.get("https://transito.mg.gov.br/")
        botaoVeiculos = navegador.find_element(By.CSS_SELECTOR, "#nav > ul > li.nav-item.yamm-fw.menu-icons-veiculos > a")
        botaoVeiculos.click()

        time.sleep(2)

        botaoImprimirCRLV = navegador.find_element(By.CSS_SELECTOR, "body > main > div > div.row > div.col-md-8 > div > div.card-body > div > div:nth-child(6) > ul > li:nth-child(2) > a > span")
        botaoImprimirCRLV.click()

        workbook = load_workbook(caminhoPlanilha)
        sheet = workbook.active

        for linhaDaVez in range(2, sheet.max_row + 1):
            placa = sheet[f'A{linhaDaVez}'].value
            renavam = sheet[f'B{linhaDaVez}'].value
            cnpj = sheet[f'C{linhaDaVez}'].value
            numeroCrv = sheet[f'D{linhaDaVez}'].value
            status = sheet[f'E{linhaDaVez}'].value

            print("Pesquisa placa: " + placa)

            if status != "OK!":
                campoPlaca = navegador.find_element(By.ID, "placa")
                campoPlaca.send_keys(placa)
                
                campoRenavam = navegador.find_element(By.ID, "renavam")
                campoRenavam.send_keys(renavam)
                
                campoCNPJ = navegador.find_element(By.ID, "cpf-cnpj")
                campoCNPJ.send_keys(cnpj)
                
                campoCRV = navegador.find_element(By.ID, "numero-crv")
                campoCRV.send_keys(numeroCrv)

                try:
                    result = solver.recaptcha(
                        sitekey='6LfVpnIUAAAAAHkISk6Z6juZcsUx6hbyJGwfnfPL',
                        url='https://transito.mg.gov.br/veiculos/documentos-de-veiculos/imprimir-crlv',
                        version='v2'
                    )
                    
                    token = result['code']
                    print('reCAPTCHA quebrado!')

                    navegador.execute_script(f"document.getElementById('g-recaptcha-response').innerHTML = '{token}'")

                    botaoPesquisa = navegador.find_element(By.CSS_SELECTOR, "#form-emitir_crlve > button")
                    navegador.execute_script("arguments[0].click();", botaoPesquisa)
                    
                    #Espera a tela do boleto aparecer certinho, esperar ele estar visivel na tela
                    telaBoleto = WebDriverWait(navegador, 60).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, "body > embed"))
                    )
            
                    time.sleep(5)
                    
                    pyautogui.hotkey('ctrl', 'p')
                    time.sleep(1.5)
                    pyautogui.hotkey('enter')
                    time.sleep(1.5)
                    print(caminho)
                    pyautogui.write(os.path.join(caminho, placa))
                    pyautogui.hotkey('enter')
                    time.sleep(2)
                    
                    sheet[f'E{linhaDaVez}'] = 'OK!'
                    workbook.save(caminhoPlanilha)
                    
                    navegador.get("https://transito.mg.gov.br/veiculos/documentos-de-veiculos/imprimir-crlv")
                    time.sleep(2)
                    
                except Exception as e:       
                    print(f"Erro ao resolver o reCAPTCHA ou imprimir: {e}")
                    time.sleep(5)

    finally:
        navegador.quit() 

# Criação da interface gráfica
root = tk.Tk()
root.title("Automação de Impressão CRLV")

# Configurações para abrir no meio da tela
root_width = 600 #largura da janela
root_height = 400 #altura da janela

screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x = (screen_width // 2) - (root_width // 2)
y = (screen_height // 2) - (root_height // 2)

root.geometry(f"{root_width}x{root_height}+{x}+{y}")
root.configure(bg='#4CAF50') 

caminhoPlanilha = ""
pathDownloads = ""

# Frame para os caminhos
frame_caminhos = tk.Frame(root, bg='#4CAF50')  # Cor de fundo do frame
frame_caminhos.pack(pady=10)

# Labels e botões organizados usando grid no frame
label_planilha = tk.Label(frame_caminhos, text="Caminho da Planilha: ", bg='#4CAF50', fg='white', font=('Helvetica', 12, 'bold'))
label_planilha.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="w")

button_planilha = tk.Button(frame_caminhos, text="Selecionar Planilha", command=selecionar_planilha, width=20, bg='white', fg='black', font=('Helvetica', 10))
button_planilha.grid(row=0, column=1, padx=5, pady=(0, 10))

label_processados = tk.Label(frame_caminhos, text="Caminho dos Processados: ", bg='#4CAF50', fg='white', font=('Helvetica', 12, 'bold'))
label_processados.grid(row=1, column=0, padx=5, pady=(0, 5), sticky="w")

button_processados = tk.Button(frame_caminhos, text="Selecionar Processados", command=selecionar_processados, width=20, bg='white', fg='black', font=('Helvetica', 10))
button_processados.grid(row=1, column=1, padx=5, pady=(0, 10))

# Botão para executar a automação
button_executar = tk.Button(root, text="Executar Automação", command=executar_automacao, width=20, bg='white', fg='black', font=('Helvetica', 10, 'bold'))
button_executar.pack(pady=20)

# Label para crédito
label_credito = tk.Label(root, text="Desenvolvido por Diogo Silva Lana", bg='#4CAF50', fg='white', font=('Helvetica', 10, 'italic'))
label_credito.pack(pady=(10, 0))  # Adiciona algum espaço acima

# Label para e-mail
label_email = tk.Label(root, text="E-mail: diogosilvalana27@gmail.com", bg='#4CAF50', fg='white', font=('Helvetica', 10))
label_email.pack()

# Label para contato
label_contato = tk.Label(root, text="Contato: (31) 99519-0505", bg='#4CAF50', fg='white', font=('Helvetica', 10))
label_contato.pack()

root.mainloop()
